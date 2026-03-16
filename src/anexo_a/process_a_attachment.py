from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from pathlib import Path

import chardet
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

try:
    from pypdf import PdfMerger
except ImportError:
    from PyPDF2 import PdfMerger


@dataclass
class ProcessamentoTXTResultado:
    arquivos_txt_processados: int
    contratos_novos: int
    ignorados_iguais: int
    duplicados_diferentes: int

    @property
    def contratos_gravados(self) -> int:
        return self.contratos_novos + self.duplicados_diferentes


@dataclass
class ValidacaoTipoResultado:
    tipo: str
    pasta: Path
    esperado: int
    encontrado: int
    faltantes: list[int]
    extras: list[int]
    extras_detalhado: list[tuple[int, str]]
    invalidos: int


class TelasPretasProcessor:
    PADRAO_BLOCO = re.compile(
        r"^@(\d+/\d+)[\s\S]*?^#\1\s*$",
        re.MULTILINE,
    )

    RE_CONTRATO_7 = re.compile(r"(?<!\d)(\d{7})(?!\d)")
    RE_CONTRATO_GERAL = re.compile(r"(?<!\d)(\d{4,15})(?!\d)")

    def __init__(
        self,
        pasta_entrada: str | Path,
        pasta_output: str | Path,
        usar_prefixo_da_subpasta: bool = False,
    ) -> None:
        self.pasta_entrada = Path(pasta_entrada)
        self.pasta_output = Path(pasta_output)
        self.usar_prefixo_da_subpasta = usar_prefixo_da_subpasta
        self.pasta_output.mkdir(parents=True, exist_ok=True)

    # =========================================================
    # UTILITÁRIOS GERAIS
    # =========================================================
    @staticmethod
    def sanitizar_nome_aba(nome: str) -> str:
        proibidos = set(r':\/?*[]')
        limpo = "".join("_" if c in proibidos else c for c in nome).strip()
        return limpo[:31] if limpo else "SEM_NOME"

    @staticmethod
    def ajustar_larguras(ws) -> None:
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 80)

    @staticmethod
    def somente_digitos(s: str) -> str:
        return "".join(ch for ch in s if ch.isdigit())

    @classmethod
    def normalizar_esperado_para_int(cls, linha: str) -> int | None:
        d = cls.somente_digitos(linha.strip())
        if not d:
            return None
        return int(d)

    @classmethod
    def extrair_numero_contrato_do_arquivo(cls, stem: str) -> int | None:
        if "_" in stem:
            _, sufixo = stem.split("_", 1)
        else:
            sufixo = stem

        sufixo = cls.somente_digitos(sufixo)
        if not sufixo:
            return None
        return int(sufixo)

    @staticmethod
    def salvar_txt_lista(caminho: Path, numeros: list[int]) -> None:
        caminho.write_text(
            "\n".join(str(n) for n in numeros) + ("\n" if numeros else ""),
            encoding="utf-8",
        )

    @staticmethod
    def achar_linha_extras(ws) -> int | None:
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip().startswith("EXTRAS"):
                return r
        return None

    @staticmethod
    def achar_pasta_origem(ws) -> Path | None:
        for r in range(1, 15):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, str) and a.strip().lower() == "pasta":
                b = ws.cell(row=r, column=2).value
                if isinstance(b, str) and b.strip():
                    return Path(b.strip())
        return None

    @staticmethod
    def achar_tipo(ws) -> str | None:
        for r in range(1, 15):
            a = ws.cell(row=r, column=1).value
            if isinstance(a, str) and a.strip().lower() == "tipo":
                b = ws.cell(row=r, column=2).value
                if isinstance(b, str) and b.strip():
                    return b.strip()
        return None

    @staticmethod
    def detectar_encoding(caminho: str | Path) -> str:
        caminho = Path(caminho)
        raw = caminho.read_bytes()
        enc = chardet.detect(raw).get("encoding") or "utf-8"
        return enc

    @staticmethod
    def escolher_fonte_para_caber(
        linhas: list[str],
        fonte_nome: str,
        largura_util: float,
        fonte_max: int = 10,
        fonte_min: int = 5,
    ) -> int:
        maior_linha = ""
        maior_largura = 0.0

        for ln in linhas:
            s = ln.rstrip("\n").rstrip("\r")
            w = stringWidth(s, fonte_nome, fonte_max)
            if w > maior_largura:
                maior_largura = w
                maior_linha = s

        if maior_largura == 0:
            return fonte_max

        tamanho = int((largura_util / maior_largura) * fonte_max)
        tamanho = max(fonte_min, min(fonte_max, tamanho))

        while tamanho > fonte_min and stringWidth(maior_linha, fonte_nome, tamanho) > largura_util:
            tamanho -= 1

        return tamanho

    @staticmethod
    def extrair_contrato_7_digitos(stem: str) -> str | None:
        m = TelasPretasProcessor.RE_CONTRATO_7.search(stem)
        return m.group(1) if m else None

    @staticmethod
    def extrair_contrato_do_nome(pdf_path: str | Path) -> str | None:
        pdf_path = Path(pdf_path)
        m = TelasPretasProcessor.RE_CONTRATO_GERAL.search(pdf_path.stem)
        return m.group(1) if m else None

    @staticmethod
    def carregar_contratos_csv(csv_path: str | Path, coluna: str = "Contrato") -> set[str]:
        csv_path = Path(csv_path)
        df = pd.read_csv(csv_path, dtype={coluna: str})

        if coluna not in df.columns:
            raise ValueError(
                f"Coluna '{coluna}' não encontrada no CSV. Colunas: {list(df.columns)}"
            )

        contratos = (
            df[coluna]
            .astype(str)
            .str.strip()
            .replace({"nan": ""})
        )
        return {c for c in contratos if c != ""}

    @staticmethod
    def _criar_logger(log_path: Path):
        def log(msg: str) -> None:
            print(msg)
            with open(log_path, "a", encoding="utf-8") as lf:
                lf.write(msg + "\n")
        return log

    def _obter_pasta_tipo(self, arquivo_entrada: Path) -> Path:
        tipo_arquivo = arquivo_entrada.stem

        if self.usar_prefixo_da_subpasta:
            rel = arquivo_entrada.relative_to(self.pasta_entrada)
            prefixo = rel.parts[0] if len(rel.parts) >= 2 else "_RAIZ"
            pasta_tipo = self.pasta_output / f"{prefixo}_{tipo_arquivo}"
        else:
            pasta_tipo = self.pasta_output / tipo_arquivo

        pasta_tipo.mkdir(parents=True, exist_ok=True)
        return pasta_tipo

    # =========================================================
    # ETAPA 1 - PROCESSAR TODOS OS TXTs DA ÁRVORE
    # =========================================================
    def processar_txts_da_arvore(self) -> ProcessamentoTXTResultado:
        total_arquivos_txt = 0
        total_contratos_novos = 0
        total_ignorados_iguais = 0
        total_duplicados_diferentes = 0

        for arquivo_entrada in self.pasta_entrada.rglob("*.txt"):
            texto = arquivo_entrada.read_text(encoding="utf-8", errors="replace")
            pasta_tipo = self._obter_pasta_tipo(arquivo_entrada)

            contador_no_txt = 0

            for bloco in self.PADRAO_BLOCO.finditer(texto):
                contrato = bloco.group(1)
                conteudo = bloco.group(0).strip() + "\n"

                nome_base = contrato.replace("/", "_")
                caminho_principal = pasta_tipo / f"{nome_base}.txt"

                if not caminho_principal.exists():
                    caminho_principal.write_text(conteudo, encoding="utf-8")
                    total_contratos_novos += 1
                    contador_no_txt += 1
                    continue

                conteudo_existente = caminho_principal.read_text(
                    encoding="utf-8",
                    errors="replace",
                )

                if conteudo_existente == conteudo:
                    total_ignorados_iguais += 1
                    continue

                i = 1
                while True:
                    caminho_dup = pasta_tipo / f"{nome_base}_DUP{i}.txt"
                    if not caminho_dup.exists():
                        caminho_dup.write_text(conteudo, encoding="utf-8")
                        total_duplicados_diferentes += 1
                        contador_no_txt += 1
                        break
                    i += 1

            print(
                f"{arquivo_entrada.name} -> {pasta_tipo.name}: "
                f"{contador_no_txt} novos/duplicados "
                f"(ignorados: dedup por igualdade)"
            )
            total_arquivos_txt += 1

        resultado = ProcessamentoTXTResultado(
            arquivos_txt_processados=total_arquivos_txt,
            contratos_novos=total_contratos_novos,
            ignorados_iguais=total_ignorados_iguais,
            duplicados_diferentes=total_duplicados_diferentes,
        )

        print("-" * 70)
        print(f"Arquivos TXT processados       : {resultado.arquivos_txt_processados}")
        print(f"Contratos gravados (novos+dup) : {resultado.contratos_gravados}")
        print(f"Novos                          : {resultado.contratos_novos}")
        print(f"Ignorados (iguais)             : {resultado.ignorados_iguais}")
        print(f"Duplicados (conteúdo diferente): {resultado.duplicados_diferentes}")
        print(f"Saída em                       : {self.pasta_output.resolve()}")

        return resultado

    # =========================================================
    # ETAPA 2 - PROCESSAR UM TXT ESPECÍFICO DE LISTAGEM
    # =========================================================
    def processar_listagem(self, arquivo_entrada: str | Path, pasta_saida: str | Path) -> int:
        arquivo_entrada = Path(arquivo_entrada)
        pasta_saida = Path(pasta_saida)
        pasta_saida.mkdir(parents=True, exist_ok=True)

        if not arquivo_entrada.exists():
            raise FileNotFoundError(
                f"Arquivo de entrada não encontrado: {arquivo_entrada.resolve()}"
            )

        texto = arquivo_entrada.read_text(encoding="utf-8", errors="replace")

        total = 0
        for bloco in self.PADRAO_BLOCO.finditer(texto):
            chave = bloco.group(1)
            conteudo = bloco.group(0).strip() + "\n"

            nome_arquivo = chave.replace("/", "_") + ".txt"
            (pasta_saida / nome_arquivo).write_text(conteudo, encoding="utf-8")
            total += 1

        print("-" * 70)
        print(f"Arquivo processado     : {arquivo_entrada.name}")
        print(f"Contratos gerados      : {total}")
        print(f"Saída em               : {pasta_saida.resolve()}")
        print("-" * 70)

        return total

    # =========================================================
    # ETAPA 3 - VALIDAR GERADOS E GERAR EXCEL
    # =========================================================
    def validar_gerados(
        self,
        txt_contratos_esperados: str | Path,
        raiz_gerados: str | Path,
        pasta_saida_validacao: str | Path,
        nome_excel: str = "validacao_por_tipo.xlsx",
    ) -> list[ValidacaoTipoResultado]:
        txt_contratos_esperados = Path(txt_contratos_esperados)
        raiz_gerados = Path(raiz_gerados)
        pasta_saida_validacao = Path(pasta_saida_validacao)

        pasta_saida_validacao.mkdir(parents=True, exist_ok=True)
        arquivo_excel = pasta_saida_validacao / nome_excel

        pasta_txt_faltantes = pasta_saida_validacao / "faltantes_txt"
        pasta_txt_faltantes.mkdir(parents=True, exist_ok=True)
        txt_faltantes_geral = pasta_txt_faltantes / "faltantes_geral.txt"

        if not txt_contratos_esperados.exists():
            raise FileNotFoundError(
                f"Arquivo de contratos esperados não encontrado: "
                f"{txt_contratos_esperados.resolve()}"
            )

        if not raiz_gerados.exists():
            raise FileNotFoundError(
                f"Pasta raiz dos gerados não encontrada: {raiz_gerados.resolve()}"
            )

        contratos_esperados: set[int] = set()
        for linha in txt_contratos_esperados.read_text(
            encoding="utf-8",
            errors="replace",
        ).splitlines():
            if not linha.strip():
                continue
            n = self.normalizar_esperado_para_int(linha)
            if n is not None:
                contratos_esperados.add(n)

        if not contratos_esperados:
            raise ValueError(
                "Nenhum contrato válido foi lido do arquivo de contratos esperados."
            )

        pastas_tipo = sorted({p.parent for p in raiz_gerados.rglob("*.txt")})
        if not pastas_tipo:
            raise ValueError(
                f"Nenhum .txt encontrado dentro de: {raiz_gerados.resolve()}"
            )

        resultado_por_tipo: list[ValidacaoTipoResultado] = []
        faltantes_geral_set: set[int] = set()

        for pasta in pastas_tipo:
            tipo = pasta.name
            encontrados_map: dict[int, list[str]] = {}
            invalidos = 0

            for arq in pasta.glob("*.txt"):
                n = self.extrair_numero_contrato_do_arquivo(arq.stem)
                if n is None:
                    invalidos += 1
                    continue
                encontrados_map.setdefault(n, []).append(arq.name)

            encontrados_set = set(encontrados_map.keys())
            faltantes = sorted(contratos_esperados - encontrados_set)
            extras = sorted(encontrados_set - contratos_esperados)

            faltantes_geral_set.update(faltantes)

            extras_detalhado: list[tuple[int, str]] = []
            for n in extras:
                for fname in sorted(encontrados_map.get(n, [])):
                    extras_detalhado.append((n, fname))

            resultado_por_tipo.append(
                ValidacaoTipoResultado(
                    tipo=tipo,
                    pasta=pasta,
                    esperado=len(contratos_esperados),
                    encontrado=len(encontrados_set),
                    faltantes=faltantes,
                    extras=extras,
                    extras_detalhado=extras_detalhado,
                    invalidos=invalidos,
                )
            )

        for r in resultado_por_tipo:
            self.salvar_txt_lista(
                pasta_txt_faltantes / f"faltantes_{r.tipo}.txt",
                r.faltantes,
            )

        self.salvar_txt_lista(txt_faltantes_geral, sorted(faltantes_geral_set))

        wb = Workbook()

        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        ws_resumo.append(
            ["Tipo", "Esperado", "Encontrado", "Faltantes", "Extras", "Inválidos", "Pasta"]
        )
        for cell in ws_resumo[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for r in resultado_por_tipo:
            ws_resumo.append(
                [
                    r.tipo,
                    r.esperado,
                    r.encontrado,
                    len(r.faltantes),
                    len(r.extras),
                    r.invalidos,
                    str(r.pasta.resolve()),
                ]
            )

        ws_resumo.freeze_panes = "A2"
        self.ajustar_larguras(ws_resumo)

        nomes_usados = set()
        for r in resultado_por_tipo:
            base = self.sanitizar_nome_aba(r.tipo)
            nome_aba = base
            i = 2
            while nome_aba in nomes_usados:
                sufixo = f"_{i}"
                nome_aba = (base[: 31 - len(sufixo)] + sufixo)[:31]
                i += 1
            nomes_usados.add(nome_aba)

            ws = wb.create_sheet(title=nome_aba)

            ws.append(["Tipo", r.tipo])
            ws.append(["Pasta", str(r.pasta.resolve())])
            ws.append(["Total esperado", r.esperado])
            ws.append(["Total encontrado", r.encontrado])
            ws.append(["Qtd faltantes", len(r.faltantes)])
            ws.append(["Qtd extras", len(r.extras)])
            ws.append(["Arquivos inválidos (nome não parseável)", r.invalidos])
            ws.append([])

            for row_cells in ws["A1":"A7"]:
                row_cells[0].font = Font(bold=True)

            ws.append(
                ["FALTANTES (contratos esperados que NÃO foram encontrados nesta pasta)"]
            )
            ws["A9"].font = Font(bold=True)
            ws.append(["Contrato (numérico)"])
            ws["A10"].font = Font(bold=True)

            row = 11
            if r.faltantes:
                for n in r.faltantes:
                    ws.cell(row=row, column=1, value=n)
                    row += 1
            else:
                ws.cell(row=row, column=1, value="(nenhum)")
                row += 1

            ws.append([])
            row += 1

            ws.append(
                ["EXTRAS (arquivos encontrados que NÃO pertencem ao lote esperado)"]
            )
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            ws.append(["Contrato extra (numérico)", "Arquivo (.txt)"])
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).font = Font(bold=True)
            row += 1

            if r.extras_detalhado:
                for n, fname in r.extras_detalhado:
                    ws.cell(row=row, column=1, value=n)
                    ws.cell(row=row, column=2, value=fname)
                    row += 1
            else:
                ws.cell(row=row, column=1, value="(nenhum)")

            ws.freeze_panes = "A11"
            self.ajustar_larguras(ws)

        wb.save(arquivo_excel)

        print(f"Excel gerado em: {arquivo_excel.resolve()}")
        print(f"TXTs de faltantes por tipo em: {pasta_txt_faltantes.resolve()}")
        print(f"TXT de faltantes geral em: {txt_faltantes_geral.resolve()}")

        return resultado_por_tipo

    # =========================================================
    # ETAPA 4 - MOVER/COPIAR EXTRAS COM BASE NO EXCEL
    # =========================================================
    def mover_extras_do_excel(
        self,
        arquivo_excel_validacao: str | Path,
        raiz_output_telaspretas: str | Path,
        mover: bool = True,
    ) -> None:
        arquivo_excel_validacao = Path(arquivo_excel_validacao)
        raiz_output_telaspretas = Path(raiz_output_telaspretas)

        pasta_extras_raiz = raiz_output_telaspretas / "extras"

        if not arquivo_excel_validacao.exists():
            raise FileNotFoundError(
                f"Excel de validação não encontrado: {arquivo_excel_validacao.resolve()}"
            )

        if not raiz_output_telaspretas.exists():
            raise FileNotFoundError(
                f"Raiz do output não encontrada: {raiz_output_telaspretas.resolve()}"
            )

        pasta_extras_raiz.mkdir(parents=True, exist_ok=True)

        wb = load_workbook(arquivo_excel_validacao, data_only=True)

        total_movidos = 0
        total_nao_encontrados = 0
        total_abas_processadas = 0

        for nome_aba in wb.sheetnames:
            if nome_aba.strip().lower() == "resumo":
                continue

            ws = wb[nome_aba]

            pasta_origem = self.achar_pasta_origem(ws)
            tipo = self.achar_tipo(ws)

            if pasta_origem is None or tipo is None:
                continue

            linha_extras = self.achar_linha_extras(ws)
            if linha_extras is None:
                continue

            row = linha_extras + 2

            destino_tipo = pasta_extras_raiz / tipo
            destino_tipo.mkdir(parents=True, exist_ok=True)

            movidos_nesta_aba = 0

            while row <= ws.max_row:
                arquivo_txt = ws.cell(row=row, column=2).value

                if arquivo_txt is None:
                    row += 1
                    continue

                if isinstance(arquivo_txt, str):
                    arquivo_txt = arquivo_txt.strip()
                    if not arquivo_txt or arquivo_txt.startswith("("):
                        break
                else:
                    break

                origem = pasta_origem / arquivo_txt
                destino = destino_tipo / arquivo_txt

                if origem.exists():
                    if mover:
                        shutil.move(str(origem), str(destino))
                    else:
                        shutil.copy2(str(origem), str(destino))
                    total_movidos += 1
                    movidos_nesta_aba += 1
                else:
                    total_nao_encontrados += 1

                row += 1

            print(f"Aba {nome_aba} (Tipo={tipo}): movidos {movidos_nesta_aba}")
            total_abas_processadas += 1

        print("-" * 70)
        print(f"Abas processadas           : {total_abas_processadas}")
        print(f"Arquivos extras movidos    : {total_movidos}")
        print(f"Extras não encontrados     : {total_nao_encontrados}")
        print(f"Destino extras             : {pasta_extras_raiz.resolve()}")
        print("-" * 70)

    # =========================================================
    # ETAPA 5 - TXT -> PDF
    # =========================================================
    def txt_para_pdf_sem_quebrar(
        self,
        txt_path: str | Path,
        pdf_path: str | Path,
    ) -> None:
        txt_path = Path(txt_path)
        pdf_path = Path(pdf_path)

        pdf_path.parent.mkdir(parents=True, exist_ok=True)

        encoding = self.detectar_encoding(txt_path)
        with open(txt_path, "r", encoding=encoding, errors="replace") as f:
            linhas = f.readlines()

        pagesize = landscape(A4)
        largura, altura = pagesize

        margem_esq = 1.5 * cm
        margem_dir = 1.5 * cm
        margem_top = 1.5 * cm
        margem_inf = 1.5 * cm

        largura_util = largura - margem_esq - margem_dir
        altura_util = altura - margem_top - margem_inf

        fonte_nome = "Courier"
        tamanho_fonte = self.escolher_fonte_para_caber(
            linhas=linhas,
            fonte_nome=fonte_nome,
            largura_util=largura_util,
            fonte_max=10,
            fonte_min=5,
        )

        leading = max(8, int(tamanho_fonte * 1.25))

        c = canvas.Canvas(str(pdf_path), pagesize=pagesize)
        c.setFont(fonte_nome, tamanho_fonte)

        x = margem_esq
        y = altura - margem_top

        linhas_por_pagina = int(altura_util // leading) or 1
        count = 0

        for ln in linhas:
            texto = ln.rstrip("\n").rstrip("\r")

            if count >= linhas_por_pagina:
                c.showPage()
                c.setFont(fonte_nome, tamanho_fonte)
                y = altura - margem_top
                count = 0

            c.drawString(x, y, texto)
            y -= leading
            count += 1

        c.save()

    def converter_pasta_txt_para_pdf(
        self,
        pasta_txt: str | Path,
        pasta_pdf: str | Path,
        recursivo: bool = True,
    ) -> None:
        pasta_txt = Path(pasta_txt).expanduser().resolve()
        pasta_pdf = Path(pasta_pdf).expanduser().resolve()
        pasta_pdf.mkdir(parents=True, exist_ok=True)

        log_path = pasta_pdf / "log.txt"
        if log_path.exists():
            log_path.unlink()

        log = self._criar_logger(log_path)

        if not pasta_txt.exists():
            raise FileNotFoundError(f"Pasta de entrada não existe: {pasta_txt}")

        padrao = "**/*.txt" if recursivo else "*.txt"
        arquivos = sorted(pasta_txt.glob(padrao))

        arquivos = [p for p in arquivos if pasta_pdf not in p.parents]

        log(f"[INFO] Pasta entrada: {pasta_txt}")
        log(f"[INFO] Pasta saída  : {pasta_pdf}")
        log(f"[INFO] TXT encontrados: {len(arquivos)} (recursivo={recursivo})")

        if not arquivos:
            log("[AVISO] Nenhum .txt encontrado. Verifique se existem .txt nessa pasta/subpastas.")
            return

        for p in arquivos[:5]:
            log(f"[AMOSTRA] {p}")

        ok = 0
        falhas = 0

        for txt in arquivos:
            try:
                rel = txt.relative_to(pasta_txt)
                pdf_dest = (pasta_pdf / rel).with_suffix(".pdf")

                self.txt_para_pdf_sem_quebrar(txt, pdf_dest)

                if not pdf_dest.exists() or pdf_dest.stat().st_size == 0:
                    raise RuntimeError("PDF não foi criado (arquivo ausente ou tamanho 0).")

                ok += 1
                log(f"[OK] {txt.name} -> {pdf_dest}")
            except Exception as e:
                falhas += 1
                log(f"[ERRO] Falha em: {txt}")
                log(f"       Motivo: {type(e).__name__}: {e}")

        log(f"[RESULTADO] PDFs gerados: {ok} | Falhas: {falhas}")
        log(f"[INFO] Log salvo em: {log_path}")

    # =========================================================
    # ETAPA 6 - MERGE GLOBAL POR CONTRATO
    # =========================================================
    def merge_global_por_contrato(
        self,
        pasta_base: str | Path,
        recursivo: bool = True,
        nome_pasta_saida: str = "merged",
    ) -> None:
        pasta_base = Path(pasta_base).expanduser().resolve()
        if not pasta_base.exists():
            raise FileNotFoundError(f"Pasta base não existe: {pasta_base}")

        pasta_merged = pasta_base / nome_pasta_saida
        pasta_merged.mkdir(parents=True, exist_ok=True)

        log_path = pasta_merged / "merge_log.txt"
        if log_path.exists():
            log_path.unlink()

        log = self._criar_logger(log_path)

        padrao = "**/*.pdf" if recursivo else "*.pdf"
        pdfs = sorted(pasta_base.glob(padrao))
        pdfs = [p for p in pdfs if pasta_merged not in p.parents]

        log(f"[INFO] Pasta base : {pasta_base}")
        log(f"[INFO] Saída      : {pasta_merged}")
        log(f"[INFO] PDFs achados: {len(pdfs)} (recursivo={recursivo})")

        if not pdfs:
            log("[AVISO] Nenhum PDF encontrado.")
            return

        grupos: dict[str, list[Path]] = {}
        sem_contrato: list[Path] = []

        for arq in pdfs:
            contrato = self.extrair_contrato_7_digitos(arq.stem)
            if not contrato:
                sem_contrato.append(arq)
                continue
            grupos.setdefault(contrato, []).append(arq)

        log(f"[INFO] Contratos identificados: {len(grupos)}")
        log(f"[INFO] PDFs sem contrato no nome: {len(sem_contrato)}")

        if sem_contrato:
            log("[AMOSTRA] Sem contrato (até 20):")
            for p in sem_contrato[:20]:
                log(f"  - {p}")

        ok = 0
        falhas = 0

        for contrato, arquivos in sorted(grupos.items(), key=lambda x: x[0]):
            try:
                arquivos = sorted(arquivos, key=lambda p: str(p).lower())
                out_pdf = pasta_merged / f"{contrato}.pdf"

                merger = PdfMerger()
                for a in arquivos:
                    merger.append(str(a))

                with open(out_pdf, "wb") as f_out:
                    merger.write(f_out)
                merger.close()

                if not out_pdf.exists() or out_pdf.stat().st_size == 0:
                    raise RuntimeError("PDF final não foi criado (ausente ou tamanho 0).")

                ok += 1
                log(f"[OK] Contrato {contrato}: {len(arquivos)} PDFs -> {out_pdf.name}")
            except Exception as e:
                falhas += 1
                log(f"[ERRO] Contrato {contrato}: {type(e).__name__}: {e}")

        log(f"[RESULTADO] Merges gerados: {ok} | Falhas: {falhas}")
        log(f"[INFO] Log salvo em: {log_path}")
        print("\n[FINALIZADO] Merge global concluído.")

    # =========================================================
    # ETAPA 7 - FILTRAR PDFs POR CSV
    # =========================================================
    def mover_pdfs_filtrados_por_csv(
        self,
        pasta_pdfs: str | Path,
        csv_path: str | Path,
        nome_pasta_saida: str = "100 maiores",
        coluna_csv: str = "Contrato",
        recursivo: bool = True,
        copiar_ao_inves_de_mover: bool = False,
    ) -> None:
        pasta_pdfs = Path(pasta_pdfs).expanduser().resolve()
        csv_path = Path(csv_path).expanduser().resolve()

        if not pasta_pdfs.exists():
            raise FileNotFoundError(f"Pasta de PDFs não existe: {pasta_pdfs}")
        if not csv_path.exists():
            raise FileNotFoundError(f"CSV não existe: {csv_path}")

        contratos_alvo = self.carregar_contratos_csv(csv_path, coluna=coluna_csv)

        pasta_saida = pasta_pdfs / nome_pasta_saida
        pasta_saida.mkdir(parents=True, exist_ok=True)

        log_path = pasta_saida / "log_filtragem.txt"
        if log_path.exists():
            log_path.unlink()

        log = self._criar_logger(log_path)

        padrao = "**/*.pdf" if recursivo else "*.pdf"
        pdfs = sorted(pasta_pdfs.glob(padrao))
        pdfs = [p for p in pdfs if pasta_saida not in p.parents]

        log(f"[INFO] Pasta PDFs: {pasta_pdfs}")
        log(f"[INFO] CSV       : {csv_path}")
        log(f"[INFO] Saída     : {pasta_saida}")
        log(f"[INFO] Contratos no CSV: {len(contratos_alvo)}")
        log(f"[INFO] PDFs encontrados: {len(pdfs)} (recursivo={recursivo})")

        movidos = 0
        ignorados_sem_contrato = 0
        ignorados_nao_alvo = 0
        falhas = 0

        for pdf in pdfs:
            try:
                contrato = self.extrair_contrato_do_nome(pdf)
                if not contrato:
                    ignorados_sem_contrato += 1
                    log(f"[SEM CONTRATO] {pdf.name}")
                    continue

                if contrato not in contratos_alvo:
                    ignorados_nao_alvo += 1
                    continue

                destino = pasta_saida / pdf.name

                if destino.exists():
                    i = 1
                    while True:
                        destino_alt = pasta_saida / f"{pdf.stem}__dup{i}{pdf.suffix}"
                        if not destino_alt.exists():
                            destino = destino_alt
                            break
                        i += 1

                if copiar_ao_inves_de_mover:
                    shutil.copy2(pdf, destino)
                    log(f"[COPIADO] {pdf.name} -> {destino.name} (contrato={contrato})")
                else:
                    shutil.move(str(pdf), str(destino))
                    log(f"[MOVIDO]  {pdf.name} -> {destino.name} (contrato={contrato})")

                movidos += 1

            except Exception as e:
                falhas += 1
                log(f"[ERRO] {pdf} | {type(e).__name__}: {e}")

        log("\n[RESULTADO]")
        log(f"  Movidos/Copiados: {movidos}")
        log(f"  Ignorados (sem contrato no nome): {ignorados_sem_contrato}")
        log(f"  Ignorados (contrato não está no CSV): {ignorados_nao_alvo}")
        log(f"  Falhas: {falhas}")
        log(f"[INFO] Log salvo em: {log_path}")

    # =========================================================
    # ETAPA 8 - RENOMEAR ARQUIVOS COM PREFIXO
    # =========================================================
    def renomear_arquivos_com_prefixo(
        self,
        pasta: str | Path,
        prefixo: str,
        recursivo: bool = False,
        apenas_pdfs: bool = True,
    ) -> None:
        pasta = Path(pasta).expanduser().resolve()

        if not pasta.exists():
            raise FileNotFoundError(f"Pasta não existe: {pasta}")

        arquivos = pasta.rglob("*") if recursivo else pasta.iterdir()

        total_renomeados = 0
        total_ignorados = 0

        for arquivo in arquivos:
            if not arquivo.is_file():
                continue

            if apenas_pdfs and arquivo.suffix.lower() != ".pdf":
                continue

            nome_atual = arquivo.name

            if nome_atual.startswith(prefixo):
                total_ignorados += 1
                continue

            novo_nome = f"{prefixo}{nome_atual}"
            novo_caminho = arquivo.with_name(novo_nome)

            arquivo.rename(novo_caminho)
            total_renomeados += 1

        print("-" * 70)
        print(f"Pasta                 : {pasta}")
        print(f"Prefixo               : {prefixo}")
        print(f"Renomeados            : {total_renomeados}")
        print(f"Ignorados (já estavam): {total_ignorados}")
        print("-" * 70)