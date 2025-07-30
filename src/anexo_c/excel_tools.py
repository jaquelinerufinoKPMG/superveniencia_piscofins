import os
import re
import shutil
import threading
import logging

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from tqdm import tqdm


class generating_excel_files:
    def __init__(self, logger_name: str = __name__):
        self.logger = logging.getLogger(logger_name)
        if not self.logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            self.logger.addHandler(handler)
            self.logger.setLevel(logging.INFO)

    def verify_and_move_file(self, file_path, invalid_dir):
        # Nome do arquivo sem a extensão
        file_name = os.path.splitext(os.path.basename(file_path))[0][1:].lstrip('0')

        # Leia o arquivo xlsx
        df = pd.read_excel(file_path, header=None)
        
        if df.iloc[1, 2] != file_name:
            # Crie o diretório se não existir
            if not os.path.exists(invalid_dir):
                os.makedirs(invalid_dir)
            # Mova o arquivo para o diretório inválido
            shutil.move(file_path, os.path.join(invalid_dir, os.path.basename(file_path)))

    def empty_row(self, worksheet, row_index):
        if worksheet.row_dimensions[row_index].hidden:
            for c in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_index, column=c).value = None
            return True
        for c in range(1, worksheet.max_column + 1):
            val = worksheet.cell(row=row_index, column=c).value
            if val not in [None, 0, -0, 0.0, -0.0]:
                return False
        return True
    
    def empty_column(self, worksheet, col_index):
        if worksheet.column_dimensions[get_column_letter(col_index)].hidden:
            return True
        for r in range(1, worksheet.max_row + 1):
            val = worksheet.cell(row=r, column=col_index).value
            if val not in [None, 0, -0, 0.0, -0.0]:
                return False
        return True
    
    def process_file(self, file_path):
        try:
            workbook = load_workbook(file_path)
            for nome_sheet in workbook.sheetnames:
                sheet = workbook[nome_sheet]
                empty_row_count = 0
                for r in range(sheet.max_row, 0, -1):
                    if self.empty_row(sheet, r):
                        empty_row_count += 1
                        if empty_row_count > 2:
                            sheet.delete_rows(r)
                    else:
                        empty_row_count = 0

                empty_column_count = 0
                for c in range(sheet.max_column, 0, -1):
                    if self.empty_column(sheet, c):
                        empty_column_count += 1
                        if empty_column_count > 2:
                            sheet.delete_cols(c)
                    else:
                        empty_column_count = 0

                for row in range(1, sheet.max_row + 1):
                    sheet.row_dimensions[row].hidden = False
                for col in range(1, sheet.max_column + 1):
                    sheet.column_dimensions[get_column_letter(col)].hidden = False

            workbook.save(file_path)
        except Exception as e:
            self.logger.error(f"Erro ao processar o arquivo {file_path}: {e}")

    def save_file(self, input_file, output_path, template_file= "template.xlsx", sheet_name="IR_CS_ANUAL"):
        output_file = f"`{output_path}/{os.path.basename(input_file)}"
        if os.path.isfile(output_file):
            return
        # Passo 1: Criar uma cópia do template
        shutil.copy(template_file, output_file)

        # Abrir o arquivo de entrada e o arquivo de saída
        workbook_input = openpyxl.load_workbook(input_file,read_only=True, data_only=True)
        workbook_output = openpyxl.load_workbook(output_file)

        # Verificar se a aba IR_CS_ANUAL existe no arquivo de entrada
        if sheet_name not in workbook_input.sheetnames:
            raise ValueError(f"A aba {sheet_name} não existe no arquivo de entrada.")

        # Selecionar a aba IR_CS_ANUAL do arquivo de entrada e de saída
        worksheet_input = workbook_input[sheet_name]
        worksheet_output = workbook_output.active
        worksheet_output.title = sheet_name

        # Copiar apenas os valores para a aba do template
        # max_row = 0
        # max_col = 0
        for row_idx, row in enumerate(worksheet_input.iter_rows(), start=1):
            # is_row_filled = False
            for col_idx, cell in enumerate(row, start=1):
                new_cell = worksheet_output.cell(row=row_idx, column=col_idx)
                new_cell.value = cell.value  # Copia apenas o valor da célula, sem fórmulas

                # Formatação condicional
                if isinstance(cell.value, str):
                    if re.match(r"^\d{8}\s", cell.value):  # 8 números seguidos de espaço
                        new_cell.font = Font(bold=True)
                    elif re.match(r"^\d{9,}", cell.value):  # Mais de 8 números no início
                        new_cell.font = Font(bold=False)
                        new_cell.alignment = Alignment(indent=1)

        # Salvar o arquivo modificado
        workbook_output.save(output_file)
        workbook_input.close()
        workbook_output.close()

    
    def process_files(self, dir: str, invalid_dir: str, definition, thread_limit: int = 60):
        """
        Processa arquivos em um diretório usando múltiplas threads, aplicando uma função de verificação em cada arquivo.
        
        Parâmetros:
        - diretorio: Caminho onde estão os arquivos a serem processados.
        - diretorio_invalido: Caminho para mover arquivos considerados inválidos.
        - funcao_verificacao: Função a ser executada em cada arquivo. Deve receber (caminho_arquivo, diretorio_invalido).
        - thread_limit: Número máximo de threads simultâneas. Padrão é 60.
        """
        os.makedirs(invalid_dir, exist_ok=True)
        files = os.listdir(dir)

        threads = []
        for idx, file in enumerate(tqdm(files, desc="Processando arquivos", total=len(files))):
            file_path = os.path.join(dir, file)
            thread = threading.Thread(target=definition, args=(file_path, invalid_dir))
            thread.start()
            threads.append(thread)

            if len(threads) >= thread_limit or idx == len(files) - 1:
                for thread in threads:
                    thread.join()
                threads = []

        print("Verificação concluída.")

    def remove_dot(self, dir: str):
        files = os.listdir(dir)

        # Itere sobre cada arquivo
        for file in files:
            # Verifique se o arquivo contém um ponto
            if '.' in file:
                # Encontre a posição do primeiro ponto
                dot_position = file.find('.')
                # Remova o primeiro ponto do nome do arquivo
                new_name = file[:dot_position] + file[dot_position+1:]
                try:
                    # Renomeie o arquivo
                    os.rename(os.path.join(dir, file), os.path.join(dir, new_name))
                except PermissionError:
                    self.logger.error(f"Permissão negada ao renomear o arquivo: {file}")

        print("Renomeação concluída.")
