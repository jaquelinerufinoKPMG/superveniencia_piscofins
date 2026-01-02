import os
import re
import shutil
import threading
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PyPDF2 import PdfReader
from pathlib import Path
from tqdm import tqdm
import pandas as pd
import win32com.client as win32
import win32com
from dotenv import find_dotenv
from environs import Env

env = Env()
env.read_env(find_dotenv())

OUTPUT_DIRECTORY = env('output_dir')
INVALID_DIRECTORY = env('invalid_dir')
WORKSHEET_NAME = env('anexo_c_worksheet_name', 'PIS_COFINS_ANUAL')
THREAD_LIMIT = 30

class export_pdf:
    def _init__(self):
        pass

    def _check_excel_files(self, excel_file, invalid_directory):
        p = Path(excel_file)

        # 1) Ignora qualquer coisa que não seja arquivo
        if not p.is_file():
            return

        # 2) Processa só Excel
        if p.suffix.lower() not in {".xlsx", ".xls", ".xlsm"}:
            return

        file_name = p.stem[1:].lstrip("0")

        try:
            # 3) Garante fechamento correto do Excel (importantíssimo em threads)
            with pd.ExcelFile(p, engine="openpyxl") as xls:
                df = pd.read_excel(
                    xls,
                    header=None,
                    dtype=str,
                    sheet_name="PIS_COFINS_ANUAL"
                )

        except Exception:
            # Qualquer erro de leitura = inválido
            output_dir = Path(invalid_directory)
            output_dir.mkdir(parents=True, exist_ok=True)
            shutil.move(p, output_dir / p.name)
            return

        # 4) Validação de conteúdo
        if str(df.iloc[1, 2]) != file_name and str(df.iloc[0, 2]) != file_name:
            output_dir = Path(invalid_directory)
            output_dir.mkdir(parents=True, exist_ok=True)
            shutil.move(p, output_dir / p.name)


    def _rename_files(self, directory_path):
        for file in os.listdir(directory_path):
            if '.' in file:
                file_name, ext = file.split('.', 1)
                letter = file_name[0]
                numbers = file_name[1:].zfill(7)
                new_name = f"{letter}{numbers}.{ext}"
                try:
                    os.rename(os.path.join(directory_path, file), os.path.join(directory_path, new_name))
                except Exception as e:
                    pass

    def _empty_row(self, worksheet, row_index):
        if worksheet.row_dimensions[row_index].hidden:
            for c in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_index, column=c).value = None
            return True
        return all(worksheet.cell(row=row_index, column=c).value in [None, 0, 0.0] for c in range(1, worksheet.max_column + 1))

    def _empty_col(self, worksheet, col_index):
        if worksheet.column_dimensions[get_column_letter(col_index)].hidden:
            return True
        return all(worksheet.cell(row=r, column=col_index).value in [None, 0, 0.0] for r in range(1, worksheet.max_row + 1))

    def _check_pdf(self, input_file):
        output_file = os.path.splitext(input_file)[0] + ".pdf"
        if os.path.isfile(output_file):
            try:
                with open(output_file, 'rb') as f:
                    pdf = PdfReader(f)
                    return len(pdf.pages) == 1
            except:
                return False
        return False

    def _export_pdf(self, input_file):
        # Evita gencache (que te deu dor de cabeça com gen_py)
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        output_file = os.path.splitext(input_file)[0] + ".pdf"

        wb = None
        try:
            wb = excel.Workbooks.Open(os.path.abspath(input_file))
            ws = wb.Worksheets(WORKSHEET_NAME)  # usa a aba certa

            last_column = ws.UsedRange.Columns.Count
            last_row = 1

            for row in range(1, ws.UsedRange.Rows.Count + 1):
                borders = ws.Cells(row, 4).Borders
                for i in range(5, 13):
                    if borders(i).Color != 0.0 or borders(i).LineStyle != -4142:
                        last_row = row
                        break

            for col in range(1, ws.UsedRange.Columns.Count + 1):
                borders = ws.Cells(5, col).Borders
                for i in range(5, 13):
                    if borders(i).Color != 0.0 or borders(i).LineStyle != -4142:
                        last_column = col
                        break

            def col_letra(i):
                letra = ''
                while i > 0:
                    i, r = divmod(i - 1, 26)
                    letra = chr(65 + r) + letra
                return letra

            rng = ws.Range(f"B1:{col_letra(last_column)}{last_row}")

            # --- page setup "resize" ---
            ws.PageSetup.PrintArea = rng.Address
            ws.DisplayPageBreaks = False
            ws.ResetAllPageBreaks()

            ps = ws.PageSetup
            ps.Zoom = False
            ps.FitToPagesWide = 1
            ps.FitToPagesTall = 1

            ps.Orientation = 1  # 1=Portrait
            ps.PaperSize = 9    # 8=A3 (use 9=A4 se quiser tentar)
            ps.LeftMargin   = excel.InchesToPoints(0.15)
            ps.RightMargin  = excel.InchesToPoints(0.15)
            ps.TopMargin    = excel.InchesToPoints(0.15)
            ps.BottomMargin = excel.InchesToPoints(0.15)
            ps.CenterHorizontally = True

            # Exporta SÓ a worksheet respeitando PrintArea
            ws.ExportAsFixedFormat(
                Type=0,
                Filename=os.path.abspath(output_file),
                IgnorePrintAreas=False,
                OpenAfterPublish=False
            )

        finally:
            if wb is not None:
                wb.Close(False)
            excel.Quit()

    def process_file(self):
        shutil.rmtree(os.path.join(win32com.__gen_path__), ignore_errors=True)
        
        os.makedirs(INVALID_DIRECTORY, exist_ok=True)
        
        arquivos = os.listdir(OUTPUT_DIRECTORY)
        threads = []
        print("Verificando arquivos inválidos...")
        for idx, arquivo in enumerate(tqdm(arquivos, desc="Verificando")):
            if arquivo.endswith(".pdf"):
                continue
            path = os.path.join(OUTPUT_DIRECTORY, arquivo)
            t = threading.Thread(target=self._check_excel_files, args=(path, INVALID_DIRECTORY))
            t.start()
            threads.append(t)
            if len(threads) >= THREAD_LIMIT or idx == len(arquivos) - 1:
                for t in threads:
                    t.join()
                threads = []
                
        print("Renomeando arquivos...")
        self._rename_files(OUTPUT_DIRECTORY)
        
        print("Convertendo para PDF...")
        arquivos = sorted(
            [f for f in os.listdir(OUTPUT_DIRECTORY) if f.endswith(".xlsx") and not self._check_pdf(os.path.join(OUTPUT_DIRECTORY, f))]
        )
        for arquivo in tqdm(arquivos, desc="Convertendo"):
            try:
                self._export_pdf(os.path.abspath(os.path.join(OUTPUT_DIRECTORY, arquivo)))
            except Exception as e:
                print(f"Erro ao converter {arquivo}: {e}")
        print("✅ Processo finalizado.")