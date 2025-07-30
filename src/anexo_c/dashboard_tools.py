import os
import time
import shutil
import logging
import pyautogui
import pandas as pd
import openpyxl
from PIL import ImageChops
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill, Protection, Side
from tqdm import tqdm
import re
import pygetwindow as gw
from pathlib import Path
import win32com.client as win32
import win32api
import win32gui
import win32con
from dotenv import load_dotenv

load_dotenv()

dir_input = os.getenv("input_dir")
dir_output = os.getenv("output_dir")
dir_invalid = os.getenv("invalid_dir")
dir_buttons = os.getenv("anexo_c_buttons_dir")

class processDashboard:
    def __init__(self, 
                 worksheet_name: str,
                 logger_name: str = __name__):

        self.worksheet_name = worksheet_name
        self.dir_output = os.path.join(dir_output, "Anexo C")
        self.dashboard_file = os.path.join(dir_input, "Anexo C", "Dashboard.xlsx")
        self.template_file = os.path.join(dir_input, "Anexo C", "template.xlsx")
        self.contract_not_found = os.path.join(dir_output,"Anexo C","contract_not_found.txt")
        self.buttons_dir = dir_buttons
        self.region = (79, 386, 802, 300)
        self.after_search = (393, 252)
        self.after_close = (48, 315)
        self.logger = logging.getLogger(logger_name)
        if not self.logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            self.logger.addHandler(handler)
            self.logger.setLevel(logging.INFO)

    def compare_images(self, img1, img2):
        diff = ImageChops.difference(img1, img2)
        return diff.getbbox() is not None

    def watch_changes(self, original):        
        begin = time.time()
        while time.time() - begin < 60:
            time.sleep(1)
            atual = pyautogui.screenshot(region=self.region)
            if self.compare_images(original, atual):
                return True
            self.logger.info("Aguardando alterações na imagem...")

        self.logger.info("Nenhuma alteração detectada em 60 segundos.")
        return False

    def wait_for_image(self, image_path, confidence=0.9, timeout=60):
        location = None
        for _ in range(timeout * 2):
            try:
                location = pyautogui.locateOnScreen(image_path, confidence=confidence)
                if location:
                    return location
            except pyautogui.ImageNotFoundException as e:
                # self.logger.info("Aguardando imagem {}...".format(image_path))
                pass
            except Exception as e:
                self.logger.error(f"Erro ao localizar imagem {image_path}: {e}")
            time.sleep(0.5)
        return None

    def save_file(self, contractID, worksheet_name, max_columns=40, max_rows=2000):
        output_file = os.path.join(self.dir_output, f"C{contractID}.xlsx")
        shutil.copy(self.template_file, output_file)

        workbook_input = openpyxl.load_workbook(self.dashboard_file, read_only=True, data_only=True)
        workbook_output = openpyxl.load_workbook(output_file)

        if worksheet_name not in workbook_input.sheetnames:
            raise ValueError("A aba {} nao existe no arquivo de entrada.".format(worksheet_name))

        worksheet_input = workbook_input[worksheet_name]
        worksheet_output = workbook_output.active
        worksheet_output.title = worksheet_name

        for row_idx, row in enumerate(worksheet_input.iter_rows(), start=1):
            if row_idx > max_rows:
                break

            for col_idx, cell in enumerate(row, start=1):
                if col_idx > max_columns:
                    break
                new_cell = worksheet_output.cell(row=row_idx, column=col_idx)
                new_cell.value = cell.value
                if isinstance(cell.value, str):
                    if re.match(r"^\d{8}\s", cell.value):
                        new_cell.font = Font(bold=True)
                    elif re.match(r"^\d{9,}", cell.value):
                        new_cell.font = Font(bold=False)
                        new_cell.alignment = Alignment(indent=1)

        workbook_output.save(output_file)
        workbook_input.close()
        workbook_output.close()
        logging.info(f"Arquivo salvo como {output_file}")

    def processed_contracts(self, file_path, file_name):
        pattern = re.compile(r'\d+')
        
        contracts = []

        for nome_arquivo in os.listdir(file_path):
            match = pattern.findall(nome_arquivo)
            if match:
                contracts.append(''.join(match))
                
        with open(file_name, 'w', encoding='utf-8') as f:
            for seq in contracts:
                f.write(seq + '\n')
    
    def process_contracts(self, contracts, monitor_index=1):

        # Obtém as informações de todos monitores conectados
        monitors = win32api.EnumDisplayMonitors(None, None)
        if monitor_index < 0 or monitor_index >= len(monitors):
            raise ValueError(f"Monitor índice {monitor_index} inválido. Existem {len(monitors)} monitores.")
        # Cada monitor é uma tupla (handle, hdc, rect)
        _, _, rect = monitors[monitor_index]
        mon_left, mon_top, mon_right, mon_bottom = rect
        mon_width = mon_right - mon_left
        mon_height = mon_bottom - mon_top

        for contract in tqdm(contracts):
            # Inicia o Excel via COM
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = True  # Torna visível para podermos mover e mostrar a janela
            excel.DisplayAlerts = False
            try:
                # Abre o Dashboard no Excel
                workbook_dashboard = excel.Workbooks.Open(self.dashboard_file)
            except Exception as e:
                self.logger.error(f"Erro ao abrir o Dashboard no Excel: {e}")
                excel.Quit()
                continue

            # Pega handle da janela do Excel
            try:
                hwnd = excel.Hwnd
                # Garante que a janela está restaurada (não minimizada/maximizada)
                win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                # Move e redimensiona para ocupar todo o monitor
                win32gui.MoveWindow(hwnd,
                                     mon_left, mon_top,
                                     mon_width, mon_height,
                                     True)
                # Traz para frente
                win32gui.SetForegroundWindow(hwnd)
            except Exception as e:
                self.logger.warning(f"Falha ao posicionar ou mostrar janela no monitor: {e}")

            # Interações via PyAutoGUI
            pyautogui.moveTo(self.after_search, duration=1)
            pyautogui.click(self.after_search)
            pyautogui.hotkey("ctrl", "home")
            # Pula se já processado
            if os.path.isfile(f"{self.dir_output}/C{contract}.xlsx") or os.path.isfile(f"{self.dir_output}/C{str(contract).zfill(7)}.xlsx"):
                workbook_dashboard.Close(SaveChanges=False)
                excel.Quit()
                continue
            self.logger.info(f"Processando contrato: {contract}")

            # Localiza dropdown e prossegue
            dropdown = self.wait_for_image('src/anexo_c/botoes/dropdown_2.png', confidence=0.9)
            if dropdown is None:
                workbook_dashboard.Close(SaveChanges=False)
                excel.Quit()
                continue
            pyautogui.click(dropdown)

            original = pyautogui.screenshot(region=self.region)

            # Marca tempo de modificação
            try:
                old_mod_time = os.path.getmtime(self.dashboard_file)
            except Exception as e:
                self.logger.error(f"Erro ao acessar 'Dashboard.xlsx': {e}")
                workbook_dashboard.Close(SaveChanges=False)
                excel.Quit()
                continue

            busca = self.wait_for_image('src/anexo_c/botoes/busca_2.png', confidence=0.9)
            if busca is None:
                workbook_dashboard.Close(SaveChanges=False)
                excel.Quit()
                continue
            pyautogui.click(busca, clicks=2)
            time.sleep(1)
            pyautogui.click(busca, clicks=2)
            pyautogui.write(str(contract), interval=0.2)
            pyautogui.moveTo(self.after_search)
            time.sleep(1)
            pyautogui.click(self.after_search)

            not_found = self.wait_for_image('src/anexo_c/botoes/not_found_2.png', confidence=0.9, timeout=3)
            if not_found:
                pyautogui.click(self.after_close)
                with open(self.contract_not_found, "a") as f:
                    f.write(f"{contract}\n")
                workbook_dashboard.Close(SaveChanges=False)
                excel.Quit()
                continue

            ok = self.wait_for_image('src/anexo_c/botoes/ok_2.png', confidence=0.9)
            if not ok:
                pyautogui.click(self.after_close)
                workbook_dashboard.Close(SaveChanges=False)
                excel.Quit()
                continue
            pyautogui.click(ok)

            if not self.watch_changes(original):
                workbook_dashboard.Close(SaveChanges=False)
                excel.Quit()
                continue

            salvar = self.wait_for_image('src/anexo_c/botoes/salvar_2.png', confidence=0.9)
            if not salvar:
                workbook_dashboard.Close(SaveChanges=False)
                excel.Quit()
                continue
            time.sleep(1)
            pyautogui.click(salvar)
            pyautogui.click(salvar)

            self.logger.info(f"Salvando dashboard para o contrato {contract}")
            # Aguarda atualização no arquivo
            for _ in range(60):
                try:
                    new_mod_time = os.path.getmtime(self.dashboard_file)
                except Exception:
                    break
                if new_mod_time > old_mod_time:
                    print("O arquivo 'Dashboard.xlsx' foi atualizado.")
                    break
                time.sleep(1)

            print("Chamando save_file")

            # Fecha e salva via COM antes de prosseguir
            try:
                workbook_dashboard.Close(SaveChanges=True)
                excel.Quit()
            except Exception as e:
                self.logger.warning(f"Não foi possível fechar o Excel: {e}")

            # Chama save_file para criar cópia tratada
            self.save_file(str(contract), self.worksheet_name)
