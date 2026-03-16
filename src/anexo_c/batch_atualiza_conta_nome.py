from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def normalize_conta(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    return re.sub(r"\s+", " ", str(x).strip())


def split_conta_nome(s: str) -> tuple[str, str]:
    if s is None:
        return "", ""
    s = str(s).strip()

    if " - " in s:
        conta, nome = s.split(" - ", 1)
        return conta.strip(), nome.strip()

    parts = re.split(r"\s*-\s*", s, maxsplit=1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()

    return s.strip(), ""


def carregar_mapa_depara(
    arquivo_de_para: str | Path,
    col_conta: str = "Conta",
    col_nome: str = "Nome",
) -> dict[str, str]:
    df = pd.read_excel(arquivo_de_para, dtype=str, sheet_name="NomeConta")
    if col_conta not in df.columns or col_nome not in df.columns:
        raise ValueError(
            f"De-para precisa ter colunas '{col_conta}' e '{col_nome}'. "
            f"Encontrei: {list(df.columns)}"
        )

    df[col_conta] = df[col_conta].map(normalize_conta)
    df[col_nome] = df[col_nome].fillna("").astype(str).str.strip()

    df = df[df[col_conta] != ""].drop_duplicates(subset=[col_conta], keep="last")
    return dict(zip(df[col_conta], df[col_nome]))


def achar_coluna_por_nome(ws, nome_coluna: str) -> int:
    """
    Procura a coluna pelo cabeçalho na 1ª linha e retorna o índice (1-based).
    """
    header_row = 1
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is not None and str(v).strip() == nome_coluna:
            return col
    raise ValueError(f"Coluna '{nome_coluna}' não encontrada na linha 1 (cabeçalho).")


def atualizar_um_arquivo(
    arquivo_excel: Path,
    mapa: dict[str, str],
    sheet_dados: str = "Dados",
    col_conta_nome: str = "Conta_Nome",
) -> tuple[int, int]:
    """
    Atualiza in-place (em memória) a coluna Conta_Nome na sheet Dados.
    Retorna (linhas_lidas, linhas_alteradas).
    """
    wb = load_workbook(arquivo_excel)
    if sheet_dados not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_dados}' não existe em {arquivo_excel.name}")

    ws = wb[sheet_dados]
    col_idx = achar_coluna_por_nome(ws, col_conta_nome)

    total = 0
    alteradas = 0

    # percorre da linha 2 até a última (assumindo cabeçalho na linha 1)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        v = cell.value

        if v is None or str(v).strip() == "":
            continue

        total += 1
        conta, nome_antigo = split_conta_nome(str(v))
        conta_key = normalize_conta(conta)

        nome_novo = mapa.get(conta_key)
        if not nome_novo or str(nome_novo).strip() == "":
            continue  # não achou no de-para, mantém como está

        novo_valor = f"{conta_key} - {str(nome_novo).strip()}"
        if str(cell.value).strip() != novo_valor:
            cell.value = novo_valor
            alteradas += 1

    return total, alteradas, wb


def processar_pasta(
    pasta_entrada: str | Path,
    arquivo_de_para: str | Path,
    pasta_saida: str | Path,
    sheet_dados: str = "Dados",
    col_conta_nome: str = "Conta_Nome",
    incluir_xlsm: bool = True,
) -> None:
    pasta_entrada = Path(pasta_entrada)
    pasta_saida = Path(pasta_saida)
    pasta_saida.mkdir(parents=True, exist_ok=True)

    mapa = carregar_mapa_depara(arquivo_de_para)

    exts = [".xlsx"]
    if incluir_xlsm:
        exts.append(".xlsm")

    arquivos = [p for p in pasta_entrada.glob("*") if p.suffix.lower() in exts and p.is_file()]
    if not arquivos:
        print("Nenhum arquivo .xlsx/.xlsm encontrado na pasta:", pasta_entrada)
        return

    print(f"Encontrados {len(arquivos)} arquivo(s) para processar.")

    for arq in arquivos:
        try:
            total, alteradas, wb = atualizar_um_arquivo(
                arq, mapa, sheet_dados=sheet_dados, col_conta_nome=col_conta_nome
            )
            if alteradas > 0:
                destino = pasta_saida / arq.name
                wb.save(destino)
                wb.close()

                print(f"- {arq.name}: linhas lidas={total}, alteradas={alteradas} -> {destino}")
            else:
                print(f"- {arq.name}: linhas lidas={total}, alteradas={alteradas} -> sem alterações")
        except Exception as e:
            print(f"! ERRO em {arq.name}: {e}")


if __name__ == "__main__":
    # Exemplo:
    # python3 batch_atualiza_conta_nome.py /caminho/pasta_excels /caminho/depara.xlsx /caminho/saida
    import sys

    if len(sys.argv) < 4:
        print("Uso: python3 batch_atualiza_conta_nome.py <pasta_entrada> <depara.xlsx> <pasta_saida>")
        raise SystemExit(1)

    processar_pasta(
        pasta_entrada=sys.argv[1],
        arquivo_de_para=sys.argv[2],
        pasta_saida=sys.argv[3],
    )