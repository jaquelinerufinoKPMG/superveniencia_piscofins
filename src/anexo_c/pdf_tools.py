import os
import shutil
import logging

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PyPDF2 import PdfReader
import win32com.client as win32


class generating_pdf_files:
    def __init__(self, logger_name: str = __name__):
        self.logger = logging.getLogger(logger_name)
        if not self.logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            handler.setFormatter(formatter)
            self.logger.addHandler(handler)
            self.logger.setLevel(logging.INFO)

    def verify_and_move_file(self, file, invalid_dir):
        file_name = os.path.splitext(os.path.basename(file))[0][1:].lstrip('0')
        df = pd.read_excel(file, header=None, dtype=str)
        if str(df.iloc[1, 2]) != file_name and str(df.iloc[0, 2]) != file_name:
            os.makedirs(invalid_dir, exist_ok=True)
            shutil.move(file, os.path.join(invalid_dir, os.path.basename(file)))

    def rename_files(self, dir):
        for file in os.listdir(dir):
            if '.' in file:
                name, ext = file.split('.', 1)
                letter = name[0]
                numbers = name[1:].zfill(7)
                new_name = f"{letter}{numbers}.{ext}"
                try:
                    os.rename(os.path.join(dir, file), os.path.join(dir, new_name))
                except Exception as e:
                    self.logger.info(f"Erro ao renomear o arquivo {file}: {e}")
                    pass

    def empty_row(self, worksheet, row_index):
        if worksheet.row_dimensions[row_index].hidden:
            for c in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_index, column=c).value = None
            return True
        return all(worksheet.cell(row=row_index, column=c).value in [None, 0, 0.0] for c in range(1, worksheet.max_column + 1))

    def empty_column(self, worksheet, col_index):
        if worksheet.column_dimensions[get_column_letter(col_index)].hidden:
            return True
        return all(worksheet.cell(row=r, column=col_index).value in [None, 0, 0.0] for r in range(1, worksheet.max_row + 1))

    def process_file(self, file_path):
        try:
            workbook = load_workbook(file_path)
            for sheet in workbook.worksheets:
                if not any(sheet.row_dimensions[row].hidden for row in range(1, sheet.max_row + 1)):
                    continue

                for r in range(sheet.max_row, 0, -1):
                    if self.empty_row(sheet, r):
                        if r <= sheet.max_row - 2:
                            sheet.delete_rows(r)

                for c in range(sheet.max_column, 0, -1):
                    if self.empty_column(sheet, c):
                        if c <= sheet.max_column - 2:
                            sheet.delete_cols(c)

                for row in range(1, sheet.max_row + 1):
                    sheet.row_dimensions[row].hidden = False
                for col in range(1, sheet.max_column + 1):
                    sheet.column_dimensions[get_column_letter(col)].hidden = False

            workbook.save(file_path)
        except Exception as e:
            self.logger.error(f"Erro ao processar o arquivo {file_path}: {e}")

    def check_pdf(self, input_file):
        output_file = os.path.splitext(input_file)[0] + ".pdf"
        if os.path.isfile(output_file):
            try:
                with open(output_file, 'rb') as f:
                    pdf = PdfReader(f)
                    return len(pdf.pages) == 1
            except:
                return False
        return False

    def xlsx_to_pdf_one_page(self, input_file):
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        output_file = os.path.splitext(input_file)[0] + ".pdf"

        try:
            workbook = excel.Workbooks.Open(input_file)
        except Exception as e:
            self.logger.error(f"Erro ao abrir {input_file}: {e}")
            return

        worksheet = workbook.Worksheets(1)
        last_column = worksheet.UsedRange.Columns.Count
        last_row = 1

        for row in range(1, worksheet.UsedRange.Rows.Count + 1):
            borders = worksheet.Cells(row, 4).Borders
            for i in range(5, 13):
                if borders(i).Color != 0.0 or borders(i).LineStyle != -4142:
                    last_row = row
                    break

        for col in range(1, worksheet.UsedRange.Columns.Count + 1):
            borders = worksheet.Cells(5, col).Borders
            for i in range(5, 13):
                if borders(i).Color != 0.0 or borders(i).LineStyle != -4142:
                    last_column = col
                    break

        def column_letter(i):
            letter = ''
            while i > 0:
                i, r = divmod(i - 1, 26)
                letter = chr(65 + r) + letter
            return letter

        worksheet.PageSetup.PrintArea = worksheet.Range(f"B1:{column_letter(last_column)}{last_row}").Address
        worksheet.PageSetup.Zoom = False
        worksheet.PageSetup.FitToPagesWide = 1
        worksheet.PageSetup.FitToPagesTall = 1

        workbook.ExportAsFixedFormat(0, output_file)
        workbook.Close(False)
        excel.Quit()