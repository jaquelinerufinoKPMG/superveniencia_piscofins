from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook
import pandas as pd

import unicodedata
import re
import time
from datetime import timedelta
import csv
import os
import zipfile
from concurrent.futures import ProcessPoolExecutor, as_completed

from tqdm import tqdm


# =========================
# CONFIG
# =========================
DATA_DIR = Path(r"data/output")          # onde estão os .xlsx
CHECKPOINT_PATH = Path("checkpoint.csv")
RESULT_PATH = Path("resultado.csv")

TARGET_TEXT = "Diferença na Base de Cálculo entre IR e CS"

# Performance / segurança
USE_PARALLEL = True                      # True = multiprocesso por arquivo
MAX_WORKERS = max(1, (os.cpu_count() or 4) - 1)  # -1 pra sobrar ar
CHECKPOINT_FLUSH_EVERY = 500             # grava a cada N resultados
DO_FSYNC = False                         # True = mais seguro, mas pode ficar bem mais lento


CHECKPOINT_FIELDS = ["arquivo", "arquivo_completo", "aba", "valor_ultimo_na_linha", "status"]


# =========================
# UTIL
# =========================
def normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))  # remove acentos
    s = re.sub(r"\s+", " ", s)  # colapsa espaços
    return s


TARGET_NORM = normalize_text(TARGET_TEXT)


def is_empty(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def last_non_empty_in_values(row_values) -> object:
    # row_values é uma tupla (values_only=True)
    for v in reversed(row_values):
        if not is_empty(v):
            return v
    return None


class RunTimer:
    """Timer simples usando perf_counter, com formatação HH:MM:SS."""
    def __init__(self):
        self.t0 = None
        self.t1 = None

    def __enter__(self):
        self.t0 = time.perf_counter()
        return self

    def __exit__(self, exc_type, exc, tb):
        self.t1 = time.perf_counter()

    @property
    def elapsed_seconds(self) -> float:
        if self.t0 is None:
            return 0.0
        end = self.t1 if self.t1 is not None else time.perf_counter()
        return end - self.t0

    def fmt(self) -> str:
        return str(timedelta(seconds=int(self.elapsed_seconds)))


def human_rate(done: int, seconds: float) -> str:
    if seconds <= 0 or done <= 0:
        return "n/a"
    r = done / seconds
    return f"{r:.2f} arquivos/s ({r*60:.0f} arquivos/min)"


# =========================
# CHECKPOINT
# =========================
def ensure_checkpoint_header(checkpoint_path: Path):
    if not checkpoint_path.exists():
        checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
        with checkpoint_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=CHECKPOINT_FIELDS)
            writer.writeheader()
            f.flush()
            if DO_FSYNC:
                os.fsync(f.fileno())


def load_checkpoint_processed_set(checkpoint_path: Path) -> set[str]:
    if not checkpoint_path.exists():
        return set()

    processed = set()
    try:
        df_ck = pd.read_csv(checkpoint_path, dtype={"arquivo": str}, encoding="utf-8-sig")
        if "arquivo" in df_ck.columns:
            processed = set(df_ck["arquivo"].dropna().astype(str).tolist())
    except Exception:
        with checkpoint_path.open("r", newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get("arquivo")
                if name:
                    processed.add(str(name))
    return processed


def append_checkpoint_rows(checkpoint_path: Path, rows: list[dict]):
    if not rows:
        return

    with checkpoint_path.open("a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CHECKPOINT_FIELDS)
        for r in rows:
            writer.writerow({
                "arquivo": r.get("arquivo"),
                "arquivo_completo": r.get("arquivo_completo"),
                "aba": r.get("aba"),
                "valor_ultimo_na_linha": r.get("valor_ultimo_na_linha"),
                "status": r.get("status"),
            })
        f.flush()
        if DO_FSYNC:
            os.fsync(f.fileno())


# =========================
# CORE: PROCESSAR 1 XLSX
# =========================
def process_one_file(xlsx_path: Path) -> dict:
    """
    Processa um único arquivo XLSX:
    - abre em read_only + data_only
    - varre sheets
    - varre linhas em streaming (iter_rows values_only)
    - compara TARGET na coluna B
    - achou => pega último valor não vazio da linha
    """
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except zipfile.BadZipFile:
        return {
            "arquivo": xlsx_path.name,
            "arquivo_completo": str(xlsx_path),
            "aba": None,
            "valor_ultimo_na_linha": None,
            "status": "erro: BadZipFile (xlsx inválido/corrompido)",
        }
    except PermissionError:
        return {
            "arquivo": xlsx_path.name,
            "arquivo_completo": str(xlsx_path),
            "aba": None,
            "valor_ultimo_na_linha": None,
            "status": "erro: PermissionError (sem acesso ao arquivo)",
        }
    except Exception as e:
        return {
            "arquivo": xlsx_path.name,
            "arquivo_completo": str(xlsx_path),
            "aba": None,
            "valor_ultimo_na_linha": None,
            "status": f"erro: {type(e).__name__} (falha ao abrir)",
        }

    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                # coluna B = índice 1
                b_val = row[1] if len(row) > 1 else None
                if b_val is None:
                    continue

                # normaliza apenas quando existe algo
                if normalize_text(b_val) == TARGET_NORM:
                    ultimo = last_non_empty_in_values(row)
                    return {
                        "arquivo": xlsx_path.name,
                        "arquivo_completo": str(xlsx_path),
                        "aba": ws.title,
                        "valor_ultimo_na_linha": ultimo,
                        "status": "ok",
                    }

        return {
            "arquivo": xlsx_path.name,
            "arquivo_completo": str(xlsx_path),
            "aba": None,
            "valor_ultimo_na_linha": None,
            "status": "nao_encontrado",
        }

    except Exception as e:
        return {
            "arquivo": xlsx_path.name,
            "arquivo_completo": str(xlsx_path),
            "aba": None,
            "valor_ultimo_na_linha": None,
            "status": f"erro: {type(e).__name__} (falha ao ler)",
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


# =========================
# MAIN PIPELINE
# =========================
def build_result_from_checkpoint():
    df_all = pd.read_csv(CHECKPOINT_PATH, encoding="utf-8-sig")
    df_all[["arquivo", "valor_ultimo_na_linha", "status"]].to_csv(
        RESULT_PATH, index=False, encoding="utf-8-sig"
    )
    return df_all


def main():
    if not DATA_DIR.exists() or not DATA_DIR.is_dir():
        raise FileNotFoundError(
            f"Pasta não encontrada: {DATA_DIR}. Ajuste DATA_DIR e coloque os .xlsx lá."
        )

    files = sorted(DATA_DIR.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Nenhum .xlsx encontrado em: {DATA_DIR}")

    ensure_checkpoint_header(CHECKPOINT_PATH)
    processed = load_checkpoint_processed_set(CHECKPOINT_PATH)
    files_to_process = [p for p in files if p.name not in processed]

    print(f"Diretório: {DATA_DIR.resolve()}")
    print(f"Total .xlsx encontrados: {len(files)}")
    print(f"Já processados (checkpoint): {len(processed)}")
    print(f"Restantes para processar: {len(files_to_process)}")
    print(f"Modo: {'PARALELO' if USE_PARALLEL else 'SEQUENCIAL'}"
          f"{f' (workers={MAX_WORKERS})' if USE_PARALLEL else ''}")
    print(f"Checkpoint flush a cada: {CHECKPOINT_FLUSH_EVERY} | fsync: {DO_FSYNC}")

    if not files_to_process:
        print("Nada a processar. Gerando resultado.csv a partir do checkpoint.csv...")
        df_all = build_result_from_checkpoint()
        print(f"OK! Gerado: {RESULT_PATH.resolve()}")
        print(df_all["status"].value_counts(dropna=False))
        return

    buffer_rows: list[dict] = []
    newly_processed = 0

    with RunTimer() as rt:
        if USE_PARALLEL:
            with ProcessPoolExecutor(max_workers=MAX_WORKERS) as ex:
                futures = [ex.submit(process_one_file, p) for p in files_to_process]

                for fut in tqdm(
                    as_completed(futures),
                    total=len(futures),
                    desc="Processando XLSX",
                    unit="arquivo",
                ):
                    r = fut.result()
                    buffer_rows.append(r)
                    newly_processed += 1

                    if len(buffer_rows) >= CHECKPOINT_FLUSH_EVERY:
                        append_checkpoint_rows(CHECKPOINT_PATH, buffer_rows)
                        buffer_rows.clear()
        else:
            for p in tqdm(files_to_process, total=len(files_to_process), desc="Processando XLSX", unit="arquivo"):
                r = process_one_file(p)
                buffer_rows.append(r)
                newly_processed += 1

                if len(buffer_rows) >= CHECKPOINT_FLUSH_EVERY:
                    append_checkpoint_rows(CHECKPOINT_PATH, buffer_rows)
                    buffer_rows.clear()

        # grava o resto
        append_checkpoint_rows(CHECKPOINT_PATH, buffer_rows)
        buffer_rows.clear()

    elapsed = rt.elapsed_seconds
    df_all = build_result_from_checkpoint()

    ok = int((df_all["status"] == "ok").sum())
    nao = int((df_all["status"] == "nao_encontrado").sum())
    err = int((df_all["status"].astype(str).str.startswith("erro")).sum())

    print(f"\nOK! Gerado: {RESULT_PATH.resolve()}")
    print(df_all["status"].value_counts(dropna=False))

    print("\n===== RESUMO DA EXECUÇÃO =====")
    print(f"Arquivos processados nesta execução: {newly_processed}")
    print(f"Total acumulado no checkpoint: {len(df_all)}")
    print(f"OK: {ok} | Não encontrado: {nao} | Erro: {err}")
    print(f"Tempo desta execução: {rt.fmt()} ({elapsed:.2f}s)")
    print(f"Throughput médio (nesta execução): {human_rate(max(newly_processed, 1), elapsed)}")
    print("==============================\n")


if __name__ == "__main__":
    main()