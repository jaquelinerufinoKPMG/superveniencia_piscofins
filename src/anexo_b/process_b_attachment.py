from tkinter import font
from src.utils.normalize_text import *
import glob
import io
import numpy as np
import os
import pandas as pd
import PyPDF2
import re
import shutil
import sys
import win32com.client as win32

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.cell.cell import MergedCell
from pandas import Timestamp
from pathlib import Path
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from tqdm import tqdm

from src.utils.normalize_text import DocumentFormatter

STATUS_COLS = [
    "Contrato",
    "Cliente - Status",
    "CNPJ/CPF - Status",
    "Valor - Status",
    "Data do Contrato - Status",
    "Data de Liquidação - Status",
    "Quantidade de Parcelas - Status",
]


class anexo_b:
    def __init__(self):
        pass

    def _create_excel_model(
        self,
        file_name: str = None,
        header_color: str = "FFFF00",
        header_text: str = "Cabeçalho 1",
        header_list: list = None,
        alignments: list = None,
        fill_colors: list = None,
        merge_columns: list = None,
        font: str = "Calibri",
        font_size: int = 10,
    ) -> Workbook:
        # Garantir que os cabeçalhos foram fornecidos
        if header_list is None:
            raise ValueError("A lista de nomes para o cabeçalho 2 deve ser fornecida.")
        if fill_colors is None:
            fill_colors = ["FFFFFF"] * len(header_list)  # Branco padrão
        if alignments is None:
            alignments = ["center"] * len(header_list)  # Alinhamento central padrão
        if merge_columns is None:
            merge_columns = [1] * len(
                header_list
            )  # Cada cabeçalho ocupa uma célula por padrão

        # Validar tamanhos das listas
        if not (
            len(header_list)
            == len(fill_colors)
            == len(alignments)
            == len(merge_columns)
        ):
            raise ValueError(
                "As listas de alignments, cores de preenchimento e merge_columns devem ter o mesmo comprimento do cabeçalho 2."
            )

        # Criar um novo workbook e planilha
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Modelo"

        # Configurar Cabeçalho 1 (linha 1)
        worksheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=sum(merge_columns)
        )
        header_cell_1 = worksheet.cell(row=1, column=1, value=header_text)
        header_cell_1.fill = PatternFill(
            start_color=header_color, end_color=header_color, fill_type="solid"
        )
        header_cell_1.font = Font(
            bold=True, name=font, size=font_size, color="FFFFFF"
        )  # Texto branco obrigatório
        header_cell_1.alignment = Alignment(horizontal="center", vertical="center")

        # Linha 2 (linha oculta)
        worksheet.row_dimensions[2].hidden = True  # Ocultar a linha 2

        # Configurar Cabeçalho 2 (linha 3)
        col_idx = 1
        for name, fill_color, alignment, merge in zip(
            header_list, fill_colors, alignments, merge_columns
        ):
            # Mesclar as colunas, se necessário
            if merge > 1:
                worksheet.merge_cells(
                    start_row=3,
                    start_column=col_idx,
                    end_row=3,
                    end_column=col_idx + merge - 1,
                )

            cell = worksheet.cell(row=3, column=col_idx, value=name)
            cell.font = Font(
                color="FFFFFF", bold=True, name=font, size=font_size
            )  # Texto branco obrigatório
            cell.fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal=alignment, vertical="center")

            # Ajustar largura das colunas mescladas
            for i in range(merge):
                worksheet.column_dimensions[get_column_letter(col_idx + i)].width = max(
                    len(name) + 2, 10
                )

            col_idx += merge  # Avançar para a próxima célula disponível

        # Salvar o arquivo
        return workbook

    def _fill_excel_data(
        self,
        workbook: Workbook,
        worksheet_name: str,
        df: pd.DataFrame,
        text_colors: list = None,
        bold: list = None,
        cell_borders: list = None,
        cell_alignments: list = None,
        merge_columns: list = None,
        column_types: list = None,
    ):

        # Obter a planilha ativa pelo nome
        if worksheet_name not in workbook.sheetnames:
            raise ValueError(f"A planilha '{worksheet_name}' não existe no workbook.")
        worksheet = workbook[worksheet_name]

        # Mapear os cabeçalhos 2 com base na linha 3 (após a linha oculta)
        header_2 = [
            cell.value for cell in worksheet[3]
        ]  # Os valores da linha 3 são os cabeçalhos 2
        col_map = {
            nome: idx + 1 for idx, nome in enumerate(header_2)
        }  # Mapear nome para índice de coluna

        # Obter as informações de mesclagem do cabeçalho
        merge_map = {}
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_row == 3:  # Somente os merges do cabeçalho 2
                merge_map[merged_range.min_col] = merged_range.max_col

        # Validar se todos os nomes do DataFrame estão no cabeçalho 2
        for column in df.columns:
            if column not in col_map:
                raise ValueError(
                    f"A coluna '{column}' no DataFrame não está no cabeçalho 2 do template."
                )

        # Preencher os dados com formatação
        for row_idx, row_data in enumerate(
            df.itertuples(index=False), start=4
        ):  # Dados começam na linha 4

            total_row = False  # Flag para verificar se a linha contém "Total"
            for col_idx, col_name in enumerate(df.columns):
                cell_idx = col_map[col_name]
                value = row_data[col_idx]
                cell = worksheet.cell(row=row_idx, column=cell_idx)

                # Verificar se a célula contém "Total" (independente de maiúsculas/minúsculas)
                if isinstance(value, str) and "total" == value.lower():
                    total_row = True

                # Determinar o tipo e formatar o valor
                if column_types and len(column_types) > col_idx:
                    data_type = column_types[col_idx].lower()
                    if data_type == "número" and isinstance(value, (int, float)):
                        cell.value = value
                        cell.number_format = "#,##0.00"
                    elif data_type == "moeda" and isinstance(value, (int, float)):
                        cell.value = value
                        cell.number_format = "R$ #,##0.00"
                    else:  # Default: texto
                        cell.value = str(value)
                else:
                    cell.value = value

                # Replicar mesclagem do cabeçalho, se aplicável
                if cell_idx in merge_map:
                    start_col = cell_idx
                    end_col = merge_map[cell_idx]
                    worksheet.merge_cells(
                        start_row=row_idx,
                        start_column=start_col,
                        end_row=row_idx,
                        end_column=end_col,
                    )
                    # Aplicar 'Wrap Text' à célula mesclada
                    cell = worksheet.cell(row=row_idx, column=start_col)
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

                # Aplicar formatações, se fornecidas
                font_color = (
                    text_colors[col_idx]
                    if text_colors and len(text_colors) > col_idx
                    else "000000"
                )
                is_bold = bold[col_idx] if bold and len(bold) > col_idx else False

                # Alterar a cor do texto para vermelho se o valor começar com "-"
                if (isinstance(value, str) and value.startswith("-")) or (
                    (isinstance(value, float) or isinstance(value, int)) and value < 0
                ):
                    font_color = "FF0000"

                cell_font = Font(color=font_color, bold=is_bold)
                cell.font = cell_font

                # Aplicar bordas, se fornecidas
                if (
                    cell_borders
                    and len(cell_borders) > col_idx
                    and cell_borders[col_idx]
                ):
                    dotted_border = Border(
                        bottom=Side(style="dotted"),
                        top=Side(style="dotted"),
                        left=Side(style="dotted"),
                        right=Side(style="dotted"),
                    )
                    cell.border = dotted_border

                # Aplicar alignments, se fornecidos
                alignment_style = Alignment(horizontal="center")
                if cell_alignments and len(cell_alignments) > col_idx:
                    cell_alignment = cell_alignments[col_idx].lower()
                    if cell_alignment in ["left", "right", "center"]:
                        alignment_style = Alignment(horizontal=cell_alignment)

                cell.alignment = alignment_style
            # Aplicar o estilo negrito em toda a linha, se "Total" estiver presente
            if total_row:
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = Font(bold=True, color=cell.font.color)

        # Mesclar células em colunas específicas para valores iguais
        if merge_columns:
            for column in merge_columns:
                col_idx = col_map[column]
                start_row = 4  # Dados começam na linha 4
                value_to_merge = None
                merge_start = None

                for row in range(start_row, start_row + len(df)):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value == value_to_merge:
                        # Continuar mesclagem
                        continue
                    else:
                        # Finalizar mesclagem anterior
                        if merge_start and merge_start != row - 1:
                            worksheet.merge_cells(
                                start_row=merge_start,
                                start_column=col_idx,
                                end_row=row - 1,
                                end_column=col_idx,
                            )
                            merged_cell = worksheet.cell(
                                row=merge_start, column=col_idx
                            )
                            merged_cell.alignment = Alignment(
                                horizontal="center", vertical="center", wrap_text=True
                            )
                            # Aplicar 'Wrap Text' à célula mesclada
                            # Garantir que o tipo e formato da célula mesclada seja o mesmo
                            if column_types and len(column_types) > col_idx:
                                data_type = column_types[col_idx].lower()
                                if data_type == "número":
                                    merged_cell.number_format = "#,##0.00"
                                elif data_type == "moeda":
                                    merged_cell.number_format = "R$ #,##0.00"

                        # Iniciar nova mesclagem
                        value_to_merge = cell.value
                        merge_start = row

                # Finalizar a última mesclagem
                if merge_start and merge_start != (start_row + len(df) - 1):
                    worksheet.merge_cells(
                        start_row=merge_start,
                        start_column=col_idx,
                        end_row=start_row + len(df) - 1,
                        end_column=col_idx,
                    )
                    merged_cell = worksheet.cell(row=merge_start, column=col_idx)
                    merged_cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    # Garantir que o tipo e formato da célula mesclada seja o mesmo
                    if column_types and len(column_types) > col_idx:
                        data_type = column_types[col_idx].lower()
                        if data_type == "número":
                            merged_cell.number_format = "#,##0.00"
                        elif data_type == "moeda":
                            merged_cell.number_format = "R$ #,##0.00"

    def _consolidate_workbooks_to_xlsx(
        self,
        workbook_list,
        input_file_path: str,
        output_directory: str,
        output_file_name: str,
        starting_cell: str = "D9",
    ):

        base_workbook = load_workbook(input_file_path)
        base_worksheet = base_workbook.active

        # Determinar linha e coluna inicial a partir da célula inicial
        starting_column = starting_cell[0].upper()
        starting_row = int(starting_cell[1:])
        starting_column_index = ord(starting_column) - ord("A") + 1

        current_row = starting_row  # Começa pela linha inicial especificada

        # Unir workbooks
        for workbook in workbook_list:
            worksheet = (
                workbook.active
            )  # Assume que o conteúdo está na aba ativa do workbook

            # Determinar o maior índice de coluna usado no workbook atual
            column_count = worksheet.max_column

            # Copiar visibilidade de colunas
            for column_index in range(1, column_count + 1):
                column_letter = get_column_letter(column_index)
                base_worksheet.column_dimensions[column_letter].hidden = (
                    worksheet.column_dimensions[column_letter].hidden
                )

            # Copiar visibilidade de linhas
            for row_index in range(1, worksheet.max_row + 1):
                if worksheet.row_dimensions[row_index].hidden:
                    base_worksheet.row_dimensions[
                        current_row + row_index - 1
                    ].hidden = True

            # Copiar conteúdo e estilos de células
            for row in worksheet.iter_rows(
                min_row=1, max_row=worksheet.max_row, max_col=column_count
            ):
                for column_index, cell in enumerate(row, start=starting_column_index):
                    new_cell = base_worksheet.cell(row=current_row, column=column_index)
                    new_cell.value = cell.value

                    # Copiar estilo, recriando os objetos de estilo
                    if cell.has_style:
                        new_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color,
                        )
                        new_cell.fill = PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color,
                        )
                        new_cell.border = Border(
                            left=cell.border.left,
                            right=cell.border.right,
                            top=cell.border.top,
                            bottom=cell.border.bottom,
                        )
                        new_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent,
                        )
                        new_cell.number_format = cell.number_format

                current_row += 1
            # for column_index in range(1, column_count + 1):
            #     column_letter = get_column_letter(column_index)
            #     base_worksheet.column_dimensions[column_letter].hidden = (
            #         worksheet.column_dimensions[column_letter].hidden
            #     )

            # # Copiar visibilidade de linhas
            # for row_index in range(1, worksheet.max_row + 1):
            #     if worksheet.row_dimensions[row_index].hidden:
            #         base_worksheet.row_dimensions[
            #             current_row + row_index - 1
            #         ].hidden = True
            # Copiar células mescladas
            for merged_cell_range in worksheet.merged_cells.ranges:
                # Traduzir o intervalo para a nova posição
                min_col, min_row, max_col, max_row = range_boundaries(
                    str(merged_cell_range)
                )
                min_col += starting_column_index - 1  # Ajustar para a coluna inicial
                max_col += starting_column_index - 1
                min_row += (
                    current_row - worksheet.max_row - 1
                )  # Ajustar para a linha atual
                max_row += current_row - worksheet.max_row - 1
                new_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                base_worksheet.merge_cells(new_range)
                # cell = ws.cell(row=min_row, column=min_col)
                # cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)

            # Adicionar uma linha em branco como separador
            current_row += 1

        # Adicionar borda três linhas abaixo da última linha
        final_border_row = current_row + 2
        bottom_border = Border(bottom=Side(style="thin"))
        for column_index in range(4, 10):  # Colunas C a J (índices 3 a 10)
            cell = base_worksheet.cell(row=final_border_row, column=column_index)
            cell.border = bottom_border

        left_border_style = Border(left=Side(style="thin"), bottom=Side(style="thin"))
        cell = base_worksheet.cell(row=final_border_row, column=3)
        cell.border = left_border_style

        right_border = Border(right=Side(style="thin"), bottom=Side(style="thin"))
        cell = base_worksheet.cell(row=final_border_row, column=10)
        cell.border = right_border

        thin = Side(style="thin")

        # Loop de primeira até a linha_borda (incluindo ela)
        for row in range(starting_row, final_border_row):
            # Coluna C = 3 -> borda ESQUERDA
            cell_c = base_worksheet.cell(row=row, column=3)
            cell_c.border = Border(left=thin)

            # Coluna J = 10 -> borda DIREITA
            cell_j = base_worksheet.cell(row=row, column=10)
            cell_j.border = Border(right=thin)

        # Remover formatação de todas as células abaixo da linha com a borda
        for row in base_worksheet.iter_rows(
            min_row=final_border_row + 1,
            max_row=base_worksheet.max_row,
            max_col=base_worksheet.max_column,
        ):
            for cell in row:
                cell.value = None
                cell.font = Font()  # Font padrão (reset)
                cell.fill = PatternFill()  # Sem preenchimento
                cell.border = Border()  # Sem borda
                cell.alignment = Alignment()  # Alinhamento padrão

        # Salvar o arquivo final
        final_output_path = Path.joinpath(output_directory, output_file_name)
        base_workbook.save(final_output_path)
        # print(f"\nArquivo salvo em: {caminho_final}")
        return final_output_path

    def xlsx_to_pdf_one_page(input_file: str, output_file: str):
        input_file = str(Path(input_file).resolve())
        output_file = str(Path(output_file).resolve())

        # garante pasta existe
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        # sanitiza nome (só por segurança)
        base = Path(input_file).stem
        safe_base = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", base)
        output_file = str(Path(output_file) / f"{safe_base}.pdf")

        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = None
        try:
            workbook = excel.Workbooks.Open(input_file, ReadOnly=True)
            worksheet = workbook.Worksheets(1)

            # setup de página (exemplo)
            worksheet.PageSetup.Zoom = False
            worksheet.PageSetup.FitToPagesWide = 1
            worksheet.PageSetup.FitToPagesTall = 1

            # evita PrintArea zuada
            worksheet.PageSetup.PrintArea = ""

            # se já existir, tenta remover
            if os.path.exists(output_file):
                try:
                    os.remove(output_file)
                except PermissionError:
                    # gera nome alternativo se estiver aberto
                    output_file = str(Path(output_file).with_name(safe_base + "_1.pdf"))

            workbook.ExportAsFixedFormat(
                0,
                output_file,
                Quality=0,  # xlQualityStandard
                IncludeDocProperties=True,
                IgnorePrintAreas=True,
                OpenAfterPublish=False,
            )

        except Exception as e:
            print("Erro ao converter XLSX para PDF:", e)
            raise
        finally:
            if workbook is not None:
                workbook.Close(False)
            excel.Quit()

    def processar_quadro_zero(self, df_quadro_zero, status_cols, contrato, comparativo):
        df_quadro = df_quadro_zero.T
        df_quadro.reset_index(inplace=True)
        df_quadro.columns = ["Ref.", "Base Analítica"]
        df_quadro["Contrato/Tela Sistêmica"] = ""

        conferencias = comparativo[
            comparativo["Contrato"] == contrato
        ].drop_duplicates()[status_cols]
        conferencias["Contrato"] = "Conferido"
        conferencias = conferencias.T.reset_index()
        conferencias.columns = ["Ref.", "Base Analítica x Contrato/Tela Sistêmica"]

        df_quadro["Base Analítica x Contrato/Tela Sistêmica"] = conferencias[
            "Base Analítica x Contrato/Tela Sistêmica"
        ]
        df_quadro["Anexo"] = f"A.{contrato}"

        df_quadro["Contrato/Tela Sistêmica"] = df_quadro.apply(
            lambda x: (
                x["Base Analítica"]
                if x["Base Analítica x Contrato/Tela Sistêmica"] == "Conferido"
                else ""
            ),
            axis=1,
        )
        df_quadro["Base Analítica x Contrato/Tela Sistêmica"] = df_quadro[
            "Base Analítica x Contrato/Tela Sistêmica"
        ].apply(lambda x: "Divergente" if x != "Conferido" else x)

        df_quadro["Base Analítica"] = df_quadro.apply(
            lambda x: (
                DocumentFormatter.format_values(
                    x["Base Analítica"], format_as_currency=True
                )
                if "Valor do Bem" in x["Ref."]
                else x["Base Analítica"]
            ),
            axis=1,
        )
        df_quadro["Contrato/Tela Sistêmica"] = df_quadro.apply(
            lambda x: (
                DocumentFormatter.format_values(
                    x["Contrato/Tela Sistêmica"], format_as_currency=True
                )
                if "Valor do Bem" in x["Ref."]
                else x["Contrato/Tela Sistêmica"]
            ),
            axis=1,
        )

        workbook = self._create_excel_model(
            file_name="Modelo_Com_Linha_Oculta.xlsx",
            header_color="002060",
            header_text="RESUMO DAS VALIDAÇÕES DAS OPERAÇÕES DE LEASING - ITAÚ UNIBANCO S.A.",
            header_list=[
                "Ref.",
                "Base Analítica",
                "Contrato/Tela Sistêmica",
                "Base Analítica x Contrato/Tela Sistêmica",
                "Anexo",
            ],
            alignments=[
                "center",
                "center",
                "center",
                "center",
                "center",
            ],  # alignments específicos por coluna
            fill_colors=["002060", "002060", "002060", "002060", "002060"],
            merge_columns=[1, 1, 1, 2, 1],
            font="Calibri",
            font_size=10,
        )

        self._fill_excel_data(
            workbook,
            "Modelo",
            df=df_quadro,
            text_colors=[None, None, None, None, "FF0000"],
            bold=[False, False, False, False, True],
            cell_borders=[True, True, True, True, True],
            cell_alignments=["left", "center", "center", "right", "center"],
            merge_columns=["Anexo"],
            column_types=["texto", "texto", "texto", "texto", "texto"],
        )

        return workbook

    def processar_quadro_um(self, df, contrato):
        df_quadro = df[df["Contrato"] == contrato].copy()

        df_quadro.rename(
            columns={"Name": "Descrição", "Value": "Valor Contabilizado"}, inplace=True
        )

        df_quadro.dropna(subset=["Valor Contabilizado"], inplace=True)
        # Filtrar apenas as linhas com valores diferentes de zero, se necessário
        df_quadro = df_quadro[df_quadro["Valor Contabilizado"] != 0]
        df_quadro = df_quadro[["Descrição", "COSIF", "Valor Contabilizado"]]
        df_quadro["Anexo"] = f"B.{contrato}"

        df_quadro["Valor Contabilizado"] = df_quadro["Valor Contabilizado"].apply(
            DocumentFormatter.format_values
        )

        df_quadro = df_quadro.sort_values(by=["COSIF"])

        workbook = self._create_excel_model(
            file_name="Modelo_Com_Linha_Oculta.xlsx",
            header_color="002060",
            header_text=f"DADOS CONTRATUAIS X REGISTRO CONTÁBIL",
            header_list=["Descrição", "COSIF", "Valor Contabilizado", "Anexo"],
            alignments=[
                "center",
                "center",
                "center",
                "center",
            ],  # alignments específicos por coluna
            fill_colors=["002060", "002060", "002060", "002060"],
            merge_columns=[1, 1, 3, 1],
            font="Calibri",
            font_size=10,
        )

        self._fill_excel_data(
            workbook,
            "Modelo",
            df_quadro,
            text_colors=[None, None, None, "FF0000"],
            bold=[False, False, False, True],
            cell_borders=[True, True, True, True],
            cell_alignments=["left", "center", "right", "center"],
            merge_columns=["Anexo"],
            column_types=["texto", "texto", "moeda", "texto"],
        )

        return workbook

    def processar_quadro_dois(self, df, contrato):

        df_quadro = df[df["Contrato"] == contrato].copy()

        df_quadro.loc[:, "Anexo"] = f"B.{contrato}"
        cols_rename = {
            "Conta": "COSIF",
            "Conta - Descrição": "Descrição",
            "Valor Líquido": "Valor Líquido Contabilizado",
        }

        df_quadro = df_quadro.rename(columns=cols_rename)
        df_quadro.fillna("", inplace=True)
        df_quadro["Valor Líquido Contabilizado"] = df_quadro[
            "Valor Líquido Contabilizado"
        ].apply(DocumentFormatter.format_values)
        df_quadro = df_quadro.round(2)
        df_quadro = df_quadro.sort_values(by=["Contrato", "COSIF"])

        workbook = self._create_excel_model(
            file_name="Modelo_Com_Linha_Oculta.xlsx",
            header_color="002060",
            header_text=f"DEMONSTRAÇÃO DO VALOR ANUAL NO RESULTADO DA SUPERVENIÊNCIA E INSUFICIÊNCIA DE DEPRECIAÇÃO CONTABILIZADOS NO PERÍODO  - CONTRATO {contrato}",
            header_list=[
                "Ano",
                "COSIF",
                "Descrição",
                "Valor Líquido Contabilizado",
                "Anexo",
            ],
            alignments=[
                "center",
                "center",
                "center",
                "center",
                "center",
            ],  # alignments específicos por coluna
            fill_colors=["002060", "002060", "002060", "002060", "002060"],
            merge_columns=[1, 1, 1, 2, 1],
            font="Calibri",
            font_size=10,
        )
        df_quadro = df_quadro[
            ["Ano", "COSIF", "Descrição", "Valor Líquido Contabilizado", "Anexo"]
        ]

        self._fill_excel_data(
            workbook,
            "Modelo",
            df_quadro[
                ["Ano", "COSIF", "Descrição", "Valor Líquido Contabilizado", "Anexo"]
            ],
            text_colors=["FF0000", None, None, None, "FF0000"],
            bold=[True, False, False, False, True],
            cell_borders=[True, True, True, True, True],
            cell_alignments=["center", "center", "center", "right", "center"],
            merge_columns=["COSIF", "Descrição", "Anexo"],
            column_types=["texto", "texto", "texto", "moeda", "texto"],
        )

        return workbook

    def processar_quadro_tres(self, df, contrato):
        df_quadro = df[df["Contrato"] == contrato].copy()
        df_quadro["Anexo"] = f"C.{contrato}"
        rename_cols = {
            "Saldos Devedores": "SALDOS DEVEDORES",
            "Saldos Credores": "SALDOS CREDORES",
            "Saldo Líquido": "SALDO LÍQUIDO",
        }
        df_quadro = df_quadro.rename(columns=rename_cols)
        df_quadro["SALDOS DEVEDORES"] = df_quadro["SALDOS DEVEDORES"].apply(
            DocumentFormatter.format_values
        )
        df_quadro["SALDOS CREDORES"] = df_quadro["SALDOS CREDORES"].apply(
            DocumentFormatter.format_values
        )
        df_quadro["SALDO LÍQUIDO"] = df_quadro["SALDO LÍQUIDO"].apply(
            DocumentFormatter.format_values
        )

        df_quadro.sort_values(by=["Contrato", "Ano"], inplace=True)

        workbook = self._create_excel_model(
            file_name="Modelo_Com_Linha_Oculta.xlsx",
            header_color="002060",
            header_text=f"DEMONSTRAÇÃO DOS SALDOS DEVEDORES E CREDORES DE SUPERVENIÊNCIA E INSUFICIÊNCIA DE DEPRECIAÇÃO POR ANO RELATIVO À OPERAÇÃO DO CONTRATO DE ARRENDAMENTO NÚMERO {contrato}",
            header_list=[
                "Ano",
                "SALDOS DEVEDORES",
                "SALDOS CREDORES",
                "SALDO LÍQUIDO",
                "Anexo",
            ],
            alignments=["center", "center", "center", "center", "center"],
            fill_colors=["002060", "002060", "002060", "002060", "002060"],
            merge_columns=[1, 1, 1, 2, 1],
            font="Calibri",
            font_size=10,
        )

        self._fill_excel_data(
            workbook,
            "Modelo",
            df_quadro[
                ["Ano", "SALDOS DEVEDORES", "SALDOS CREDORES", "SALDO LÍQUIDO", "Anexo"]
            ],
            text_colors=["FF0000", None, None, None, "FF0000"],
            bold=[True, False, False, False, True],
            cell_borders=[True, True, True, True, True],
            cell_alignments=["center", "right", "right", "right", "center"],
            merge_columns=["Anexo"],
            column_types=["texto", "moeda", "moeda", "moeda", "texto"],
        )

        return workbook

    def processar_quadro_quatro(self, df, contrato):
        df_quadro = df[df["Contrato"] == contrato].copy()
        df_quadro["Anexo"] = f"C.{contrato}"
        rename_cols = {
            "Receitas": "Receita de Contraprestação - Inclui Superveniência (1)",
            "Exclusão": "Exclusão - Recuperção Baixada como Prejuízo (2)",
            "Dedução": "Dedução - Depreciação/Outras Despesas (3)",
            "Base de Cálculo": "Base de Cálculo (01)-(02)-(03)",
        }
        df_quadro = df_quadro.rename(columns=rename_cols)
        df_quadro = df_quadro.rename(columns=rename_cols)
        df_quadro["Receita de Contraprestação - Inclui Superveniência (1)"] = df_quadro[
            "Receita de Contraprestação - Inclui Superveniência (1)"
        ].apply(DocumentFormatter.format_values)
        df_quadro["Exclusão - Recuperção Baixada como Prejuízo (2)"] = df_quadro[
            "Exclusão - Recuperção Baixada como Prejuízo (2)"
        ].apply(DocumentFormatter.format_values)
        df_quadro["Dedução - Depreciação/Outras Despesas (3)"] = df_quadro[
            "Dedução - Depreciação/Outras Despesas (3)"
        ].apply(DocumentFormatter.format_values)
        df_quadro["Base de Cálculo (01)-(02)-(03)"] = df_quadro[
            "Base de Cálculo (01)-(02)-(03)"
        ].apply(DocumentFormatter.format_values)

        df_quadro.sort_values(by=["Contrato", "Ano"], inplace=True)

        workbook = self._create_excel_model(
            file_name="Modelo_Com_Linha_Oculta.xlsx",
            header_color="002060",
            header_text=f"BASE DE CÁLCULO DO PIS E DA COFINS RELATIVO À OPERAÇÃO DO CONTRATO DE ARRENDAMENTO NÚMERO {contrato}",
            header_list=[
                "Ano",
                "Receita de Contraprestação - Inclui Superveniência (1)",
                "Exclusão - Recuperção Baixada como Prejuízo (2)",
                "Dedução - Depreciação/Outras Despesas (3)",
                "Base de Cálculo (01)-(02)-(03)",
                "Anexo",
            ],
            alignments=[
                "center",
                "center",
                "center",
                "center",
                "center",
                "center",
            ],  # alignments específicos por coluna
            fill_colors=["002060", "002060", "002060", "002060", "002060", "002060"],
            merge_columns=[1, 1, 1, 1, 1, 1],
            font="Calibri",
            font_size=10,
        )
        self._fill_excel_data(
            workbook,
            "Modelo",
            df_quadro[
                [
                    "Ano",
                    "Receita de Contraprestação - Inclui Superveniência (1)",
                    "Exclusão - Recuperção Baixada como Prejuízo (2)",
                    "Dedução - Depreciação/Outras Despesas (3)",
                    "Base de Cálculo (01)-(02)-(03)",
                    "Anexo",
                ]
            ],
            text_colors=["FF0000", None, None, None, None, "FF0000"],
            bold=[True, False, False, False, False, True],
            cell_borders=[True, True, True, True, True, True],
            cell_alignments=["center", "right", "right", "right", "right", "center"],
            merge_columns=["Anexo"],
            column_types=["texto", "moeda", "moeda", "moeda", "moeda", "texto"],
        )

        return workbook

    def processar_quadro_cinco(self, df, contrato):
        df_quadro = df[df["Contrato"] == contrato].copy()
        df_quadro["Anexo"] = f"C.{contrato}"
        rename_cols = {
            "Receitas": "Receita de Contraprestação - Não Inclui Superveniência (1)",
            "Exclusão": "Exclusão - Recuperção Baixada como Prejuízo (2)",
            "Dedução": "Dedução - Depreciação/Outras Despesas (3)",
            "Base de Cálculo": "Base de Cálculo (01)-(02)-(03)",
        }
        df_quadro = df_quadro.rename(columns=rename_cols)
        df_quadro["Receita de Contraprestação - Não Inclui Superveniência (1)"] = (
            df_quadro[
                "Receita de Contraprestação - Não Inclui Superveniência (1)"
            ].apply(DocumentFormatter.format_values)
        )
        df_quadro["Exclusão - Recuperção Baixada como Prejuízo (2)"] = df_quadro[
            "Exclusão - Recuperção Baixada como Prejuízo (2)"
        ].apply(DocumentFormatter.format_values)
        df_quadro["Dedução - Depreciação/Outras Despesas (3)"] = df_quadro[
            "Dedução - Depreciação/Outras Despesas (3)"
        ].apply(DocumentFormatter.format_values)
        df_quadro["Base de Cálculo (01)-(02)-(03)"] = df_quadro[
            "Base de Cálculo (01)-(02)-(03)"
        ].apply(DocumentFormatter.format_values)

        df_quadro.sort_values(by=["Contrato", "Ano"], inplace=True)

        workbook = self._create_excel_model(
            file_name="Modelo_Com_Linha_Oculta.xlsx",
            header_color="002060",
            header_text=f"BASE DE CÁLCULO DO PIS E DA COFINS SEM O EFEITO DA SUPERVENIÊNCIA/INSUFICIÊNCIA DE DEPRECIAÇÃO RELATIVO À OPERAÇÃO DO CONTRATO DE ARRENDAMENTO NÚMERO {contrato}",
            header_list=[
                "Ano",
                "Receita de Contraprestação - Não Inclui Superveniência (1)",
                "Exclusão - Recuperção Baixada como Prejuízo (2)",
                "Dedução - Depreciação/Outras Despesas (3)",
                "Base de Cálculo (01)-(02)-(03)",
                "Anexo",
            ],
            alignments=[
                "center",
                "center",
                "center",
                "center",
                "center",
                "center",
            ],  # alignments específicos por coluna
            fill_colors=[
                "002060",
                "002060",
                "002060",
                "002060",
                "002060",
                "002060",
            ],
            merge_columns=[1, 1, 1, 1, 1, 1],
            font="Calibri",
            font_size=10,
        )
        self._fill_excel_data(
            workbook,
            "Modelo",
            df_quadro[
                [
                    "Ano",
                    "Receita de Contraprestação - Não Inclui Superveniência (1)",
                    "Exclusão - Recuperção Baixada como Prejuízo (2)",
                    "Dedução - Depreciação/Outras Despesas (3)",
                    "Base de Cálculo (01)-(02)-(03)",
                    "Anexo",
                ]
            ],
            text_colors=["FF0000", None, None, None, None, "FF0000"],
            bold=[True, False, False, False, False, True],
            cell_borders=[True, True, True, True, True, True],
            cell_alignments=["center", "right", "right", "right", "right", "center"],
            merge_columns=["Anexo"],
            column_types=["texto", "moeda", "moeda", "moeda", "moeda", "texto"],
        )

        return workbook

    def processar_quadro_seis(self, df, contrato):
        df_quadro = df[df["Contrato"] == contrato].copy()
        df_quadro["Anexo"] = f"C.{contrato}"
        rename_cols = {
            "Base Q4": "Base Com Efeito da Superveniência/Insuficiência",
            "Base Q5": "Base Sem Efeito da Superveniência/Insuficiência",
        }
        df_quadro = df_quadro.rename(columns=rename_cols)
        df_quadro["Base Com Efeito da Superveniência/Insuficiência"] = df_quadro[
            "Base Com Efeito da Superveniência/Insuficiência"
        ].apply(DocumentFormatter.format_values)
        df_quadro["Base Sem Efeito da Superveniência/Insuficiência"] = df_quadro[
            "Base Sem Efeito da Superveniência/Insuficiência"
        ].apply(DocumentFormatter.format_values)
        df_quadro["Diferença"] = df_quadro["Diferença"].apply(
            DocumentFormatter.format_values
        )

        df_quadro.sort_values(by=["Contrato", "Ano"], inplace=True)

        workbook = self._create_excel_model(
            file_name="Modelo_Com_Linha_Oculta.xlsx",
            header_color="002060",
            header_text=f"COMPARATIVO ENTRE A BASE DE CÁLCULO DO PIS E DA COFINS COM E SEM O EFEITO DA SUPERVENIÊNCIA/INSUFICIÊNCIA DE DEPRECIAÇÃO RELATIVO À OPERAÇÃO DO CONTRATO DE ARRENDAMENTO NÚMERO {contrato}",
            header_list=[
                "Ano",
                "Base Com Efeito da Superveniência/Insuficiência",
                "Base Sem Efeito da Superveniência/Insuficiência",
                "Diferença",
                "Anexo",
            ],
            alignments=[
                "center",
                "center",
                "center",
                "center",
                "center",
            ],  # alignments específicos por coluna
            fill_colors=["002060", "002060", "002060", "002060", "002060"],
            merge_columns=[1, 1, 1, 2, 1],
            font="Calibri",
            font_size=10,
        )

        self._fill_excel_data(
            workbook,
            "Modelo",
            df_quadro[
                [
                    "Ano",
                    "Base Com Efeito da Superveniência/Insuficiência",
                    "Base Sem Efeito da Superveniência/Insuficiência",
                    "Diferença",
                    "Anexo",
                ]
            ],
            text_colors=["FF0000", None, None, None, "FF0000"],
            bold=[True, False, False, False, True],
            cell_borders=[True, True, True, True, True],
            cell_alignments=["center", "right", "right", "right", "center"],
            merge_columns=["Anexo"],
            column_types=["texto", "moeda", "moeda", "moeda", "texto"],
        )

        return workbook

    def criar_folha_rosto(
        self,
        contrato,
        df_quadro_zero,
        comparativo,
        caminho_destino,
        df_quadro_um,
        df_quadro_dois,
        df_quadro_tres,
        df_quadro_quatro,
        df_quadro_cinco,
        df_quadro_seis,
        caminho_base,
        replace=False,
    ):

        nome_final = f"RESUMO_{contrato}.xlsx"

        if not replace and os.path.isfile(Path.joinpath(caminho_destino, nome_final)):
            return
        os.makedirs(caminho_destino, exist_ok=True)

        wb_resumo = self.processar_quadro_zero(
            df_quadro_zero, STATUS_COLS, contrato, comparativo
        )
        wb_quadro_um = self.processar_quadro_um(df_quadro_um, contrato)
        wb_quadro_dois = self.processar_quadro_dois(df_quadro_dois, contrato)
        wb_quadro_tres = self.processar_quadro_tres(df_quadro_tres, contrato)
        wb_quadro_quatro = self.processar_quadro_quatro(df_quadro_quatro, contrato)
        wb_quadro_cinco = self.processar_quadro_cinco(df_quadro_cinco, contrato)
        wb_quadro_seis = self.processar_quadro_seis(df_quadro_seis, contrato)

        self._consolidate_workbooks_to_xlsx(
            workbook_list=[
                wb_resumo,
                wb_quadro_um,
                wb_quadro_dois,
                wb_quadro_tres,
                wb_quadro_quatro,
                wb_quadro_cinco,
                wb_quadro_seis,
            ],
            input_file_path=caminho_base,
            output_directory=caminho_destino,
            output_file_name=nome_final,
            starting_cell="D9",
        )

    def safe_criar_folha_rosto(
        self,
        contrato,
        df_base_contrato,
        comparativo,
        caminho_destino,
        df_quadro_um,
        df_quadro_dois,
        df_quadro_tres,
        df_quadro_quatro,
        df_quadro_cinco,
        df_quadro_seis,
        caminho_base,
        replace,
    ):
        try:
            self.criar_folha_rosto(
                contrato,
                df_base_contrato,
                comparativo,
                caminho_destino,
                df_quadro_um,
                df_quadro_dois,
                df_quadro_tres,
                df_quadro_quatro,
                df_quadro_cinco,
                df_quadro_seis,
                caminho_base,
                replace,
            )
        except Exception as e:
            # imprime nome do grupo, mensagem de erro e traceback completo
            print(f"[ERRO] no grupo {contrato}: {e}", file=sys.stderr)
            # traceback.print_exc()

    def process_anexo_pdfs(self, caminho_base, limite_mb=5, contador_inicial=0):
        contratos_invalidos = []
        limite_bytes = limite_mb * 1024 * 1024
        lote = os.path.basename(os.path.dirname(caminho_base))
        anexo = os.path.basename(caminho_base)

        arquivos_info = []
        for f in os.listdir(caminho_base):
            if f.endswith(".pdf"):
                caminho_pdf = os.path.join(caminho_base, f)
                tamanho = os.path.getsize(caminho_pdf)
                if re.sub(r"\D", "", f) not in contratos_invalidos:
                    arquivos_info.append((re.sub(r"\D", "", f), caminho_pdf, tamanho))

        pasta_final = f"{caminho_base}_{lote}_partes"
        os.makedirs(pasta_final, exist_ok=True)

        def dividir_pdf(origem, tamanho_max, label):
            partes = []
            with open(origem, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                writer = None
                parte_indice = 1
                tamanho_acumulado = 0
                for page in reader.pages:
                    temp_writer = PyPDF2.PdfWriter()
                    temp_writer.add_page(page)

                    with open("temp_part.pdf", "wb") as temp_out:
                        temp_writer.write(temp_out)
                    parte_tamanho = os.path.getsize("temp_part.pdf")
                    if not writer or (tamanho_acumulado + parte_tamanho > tamanho_max):
                        if writer:
                            nome_parte = os.path.join(
                                pasta_final, f"{label}_parte_{parte_indice}.pdf"
                            )
                            with open(nome_parte, "wb") as out:
                                writer.write(out)
                            partes.append(nome_parte)
                            parte_indice += 1
                        writer = PyPDF2.PdfWriter()
                        tamanho_acumulado = 0
                    writer.add_page(page)
                    tamanho_acumulado += parte_tamanho
                if writer:
                    nome_parte = os.path.join(
                        pasta_final, f"{label}_parte_{parte_indice}.pdf"
                    )
                    with open(nome_parte, "wb") as out:
                        writer.write(out)
                    partes.append(nome_parte)
            os.remove("temp_part.pdf")
            return partes

        grupos = []
        grupo_atual = []
        tamanho_atual = 0
        for pasta, caminho_pdf, tamanho in arquivos_info:
            if tamanho > limite_bytes:
                nome_base = os.path.splitext(os.path.basename(caminho_pdf))[0]
                partes = dividir_pdf(caminho_pdf, limite_bytes, nome_base)
                for parte in partes:
                    p_size = os.path.getsize(parte)
                    parte_nome = os.path.basename(parte)
                    if p_size + tamanho_atual <= limite_bytes:
                        grupo_atual.append((pasta, parte_nome))
                        tamanho_atual += p_size
                    else:
                        if grupo_atual:
                            grupos.append(grupo_atual)
                        grupo_atual = [(pasta, parte_nome)]
                        tamanho_atual = p_size
            else:
                if tamanho + tamanho_atual <= limite_bytes:
                    grupo_atual.append((pasta, caminho_pdf))
                    tamanho_atual += tamanho
                else:
                    if grupo_atual:
                        grupos.append(grupo_atual)
                    grupo_atual = [(pasta, caminho_pdf)]
                    tamanho_atual = tamanho
        if grupo_atual:
            grupos.append(grupo_atual)

        dados_excel = []
        final_pdfs = []
        for i, grupo in enumerate(tqdm(grupos, desc="Processando grupos"), start=1):

            for p, arquivo in grupo:
                parte = (
                    f" - Parte {arquivo.split('_parte_')[-1].split('.pdf')[0]}"
                    if "_parte_" in arquivo
                    else ""
                )
                dados_excel.append(
                    {
                        "Contrato Número": p,
                        "Lote": f"{anexo}_{lote}_parte_{i}.pdf",
                        "Parte do contrato": parte,
                    }
                )

            escritor_final = PyPDF2.PdfWriter()
            for p, arquivo in grupo:
                # Cria uma página de cabeçalho com o número do contrato
                buffer = io.BytesIO()
                c = canvas.Canvas(buffer, pagesize=letter)
                c.setFont("Helvetica-Bold", 14)
                width, height = letter
                texto = f"Contrato Número: {p}"
                c.drawCentredString(width / 2, height / 2, texto)
                c.showPage()
                c.save()
                buffer.seek(0)

                leitor_temp = PyPDF2.PdfReader(buffer)
                for pagina_temp in leitor_temp.pages:
                    escritor_final.add_page(pagina_temp)

                if os.path.exists(arquivo):
                    caminho_completo = arquivo
                else:
                    caminho_completo = os.path.join(pasta_final, arquivo)

                with open(caminho_completo, "rb") as pdf_in:
                    leitor = PyPDF2.PdfReader(pdf_in)
                    for pagina in leitor.pages:
                        escritor_final.add_page(pagina)

            nome_pdf_final = os.path.join(pasta_final, f"{anexo}_{lote}_parte_{i}.pdf")
            with open(nome_pdf_final, "wb") as saida:
                escritor_final.write(saida)
            final_pdfs.append(nome_pdf_final)

        pdf_to_subfolder = {}
        max_docs = 14
        max_folder_size = 140 * 1024 * 1024  # 140 MB em bytes
        current_group_files = []
        current_group_size = 0
        folder_index = contador_inicial

        for pdf_file in final_pdfs:
            file_size = os.path.getsize(pdf_file)
            if (len(current_group_files) >= max_docs) or (
                current_group_size + file_size > max_folder_size
            ):
                subfolder = os.path.join(pasta_final, f"{folder_index}_{lote}_grupo")
                os.makedirs(subfolder, exist_ok=True)
                for f_pdf in current_group_files:
                    destino = os.path.join(subfolder, os.path.basename(f_pdf))
                    os.rename(f_pdf, destino)
                    pdf_to_subfolder[os.path.basename(f_pdf)] = (
                        f"{folder_index}_{lote}_grupo"
                    )
                folder_index += 1
                current_group_files = []
                current_group_size = 0
            current_group_files.append(pdf_file)
            current_group_size += file_size

        if current_group_files:
            subfolder = os.path.join(pasta_final, f"{folder_index}_{lote}_grupo")
            os.makedirs(subfolder, exist_ok=True)
            for f_pdf in current_group_files:
                destino = os.path.join(subfolder, os.path.basename(f_pdf))
                os.rename(f_pdf, destino)
                pdf_to_subfolder[os.path.basename(f_pdf)] = (
                    f"{folder_index}_{lote}_grupo"
                )

        df = pd.DataFrame(dados_excel)
        df["Pasta Interna"] = df["Lote"].apply(
            lambda lote: pdf_to_subfolder.get(lote, "")
        )

        df.to_excel(
            os.path.join(
                str(caminho_base) + f"_{lote}_partes", f"{anexo}_{lote}_partes.xlsx"
            ),
            index=False,
        )

        temp_files = ["temp_part.pdf"]
        for temp_file in temp_files:
            if os.path.exists(temp_file):
                os.remove(temp_file)

        for p in glob.glob(str(caminho_base) + f"_{lote}_partes/*.pdf"):
            os.remove

        shutil.make_archive(
            str(caminho_base) + f"_{lote}_partes",
            "zip",
            str(caminho_base) + f"_{lote}_partes",
        )
        shutil.rmtree(str(caminho_base) + f"_{lote}_partes")
        return folder_index + 1
