import pyautogui
import time
import xlwings as xw
import glob
import os
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font, Alignment
import shutil
import re
import pandas as pd
from PIL import ImageChops

# python -m pip install pyautogui pandas tqdm openpyxl xlwings opencv-python


def imagem_diferente(img1, img2):
    """
    Compara duas imagens e retorna True se houver diferença,
    False se forem idênticas.
    """
    # ImageChops.difference retorna uma nova imagem com a diferença absoluta
    diff = ImageChops.difference(img1, img2)
    # Se getbbox() retornar algo diferente de None, há pelo menos uma diferença.
    return diff.getbbox() is not None

def monitorar_alteracao(original):
    
    # Define a região: (x_inicial, y_inicial, largura, altura)
    # Como x1=79, y1=386 e x2=881, y2=686, a largura = 881 - 79 = 802 e a altura = 686 - 386 = 300.
 
    
    # Captura a imagem original da região

    
    inicio = time.time()
    while time.time() - inicio < 60:  # Loop de 60 segundos
        time.sleep(1)  # Espera 2 segundos
        atual = pyautogui.screenshot(region=region)
        if imagem_diferente(original, atual):
            return True  # Houve alteração na imagem
        print("monitorar_alteracao ...")
    return False  # Nenhuma alteração foi detectada em 60 segundos

def wait_for_image(image_path, confidence=0.9, timeout=60):
    """
    Aguarda até que a imagem especificada apareça na tela.
    Retorna a localização (caixa delimitadora) do elemento.
    Levanta uma exceção se o elemento não for encontrado dentro do tempo limite.
    """
    location = None
    for _ in range(timeout*2):
        try:    
            location = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if location is not None:
                return location
        except Exception as _:
            print("wait_for_image ...")
            time.sleep(0.5)
    return location
        

def is_row_empty(sheet, row_idx):
    """
    Verifica se uma linha está vazia (todas as células são None ou valores vazios).
    """
    for cell in sheet[row_idx]:
        if cell.value not in (None, ""):
            return False
    return True

def is_col_empty(sheet, col_idx):
    """
    Verifica se uma coluna está vazia (todas as células são None ou valores vazios).
    """
    for row in sheet.iter_rows():
        if row[col_idx - 1].value not in (None, ""):
            return False
    return True

def salve_file(contrato):
    input_file = r"Superveniência\superveniencia_piscofins\Input\Anexo C\Dashboard.xlsb"
    template_file = "template.xlsx"
    output_file = f"Output/C{contrato}.xlsx"

    # Passo 1: Criar uma cópia do template
    shutil.copy(template_file, output_file)

    # Abrir o arquivo de entrada e o arquivo de saída
    wb_input = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    wb_output = openpyxl.load_workbook(output_file)

    # Verificar se a aba PIS_COFINS_ANUAL existe no arquivo de entrada
    if "PIS_COFINS_ANUAL" not in wb_input.sheetnames:
        raise ValueError("A aba 'PIS_COFINS_ANUAL' não existe no arquivo de entrada.")

    # Selecionar a aba PIS_COFINS_ANUAL do arquivo de entrada e de saída
    ws_input = wb_input["PIS_COFINS_ANUAL"]
    ws_output = wb_output.active
    ws_output.title = "PIS_COFINS_ANUAL"

    # Copiar apenas os valores para a aba do template
    for row_idx, row in enumerate(ws_input.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            new_cell = ws_output.cell(row=row_idx, column=col_idx)
            new_cell.value = cell.value  # Copia apenas o valor da célula, sem fórmulas

            # Formatação condicional
            if isinstance(cell.value, str):
                if re.match(r"^\d{8}\s", cell.value):  # 8 números seguidos de espaço
                    new_cell.font = Font(bold=True)
                elif re.match(r"^\d{9,}", cell.value):  # Mais de 8 números no início
                    new_cell.font = Font(bold=False)
                    new_cell.alignment = Alignment(indent=1)

    wb_output.save(output_file)
    wb_input.close()
    wb_output.close()

    print(f"Arquivo salvo como {output_file}")

# Espera para abrir o programa

# Movimentação inicial do mouse
pyautogui.moveTo(x=10, y=10, duration=1)
# Define coordenadas de clique (exemplo)
x = 457
y = 402
x1 = 172
y1= 20
pyautogui.moveTo(x=x, y=y, duration=1)
pyautogui.click(x=x, y=y)

# Carregamento dos contratos via pandas
# datalist = glob.glob("C:/Projects/local-doc-data-miner/projects_data/Itau/OneDrive_2024-10-21/tela_preta/itens_96/*")
# contratos = [os.path.basename(path).lstrip('0') for path in datalist]

# Caminho do arquivo .xlsx
data_file = r"Superveniência\superveniencia_piscofins\Input\Anexo C\Contratos_2014.xlsx"
data = pd.read_excel(data_file)

# sort_file = "sort.xlsx"
# sort = pd.read_excel(sort_file)
contratos_validos = data["Contrato"].values.tolist()
# sort = sort[sort["n° do contrato"].isin(contratos_validos)]
# contratos = sort["n° do contrato"].values.tolist()
# contratos_validos.reverse()
contratos = contratos_validos
# contratos = list(reversed(contratos_validos))

# print(len(contratos))
# print(len(contratos) / 7)
# contratos = contratos[:7370]
# Se necessário, divida os contratos em lotes conforme as linhas comentadas:
# contratos = contratos[7370:7370*2]
# contratos = contratos[7370*2:7370*3]
# contratos = contratos[7370*3:7370*4]
# contratos = contratos[7370*4:7370*5]
# contratos = contratos[7370*5:7370*6]
# contratos = contratos[7370*6:]
# Leia o arquivo nao_encontrados_3.txt e obtenha os contratos não encontrados
input_file = r"Superveniência\superveniencia_piscofins\Input\Anexo C\Dashboard.xlsb"
nao_encontrados_3 = set()
if os.path.isfile("nao_encontrados_3.txt"):
    with open("nao_encontrados_3.txt", "r") as f:
        nao_encontrados_3 = set(line.strip() for line in f)

# Filtre os contratos removendo aqueles que estão na lista de não encontrados
contratos = [contrato for contrato in contratos if str(contrato) not in nao_encontrados_3]
for contrato in tqdm(contratos):
    if os.path.isfile(f"testes/C{contrato}.xlsx") or os.path.isfile(f"testes/C{str(contrato).zfill(7)}.xlsx"):
        continue
    print(contrato)


    # Espera e clica no botão 'dropdow'

    dropdow = wait_for_image('botoes/dropdown_2.png', confidence=0.9)
    print("dropdow encontrado")
    region = (79, 386, 802, 300)
    original = pyautogui.screenshot(region=region)
    if dropdow is None:
        continue  # Pula para o próximo contrato se o botão não for encontrado
    pyautogui.click(dropdow)
    # Aguarda o botão 'busca' e clica nele
    try:
        old_mod_time = os.path.getmtime(input_file)
    except Exception as e:
        print(f"Erro ao acessar 'Dashboard.xlsx': {e}")
        continue
    busca = wait_for_image('botoes/busca_2.png', confidence=0.9)
    print("busca encontrado")
    if busca is None:
        continue 
    pyautogui.click(busca, clicks=2)
    time.sleep(1)  # Pequena pausa para garantir que o campo esteja ativo
    pyautogui.click(busca)
    pyautogui.click(busca)
    # time.sleep(5)


    # Digita o número do contrato
    pyautogui.write(str(contrato), interval=0.2)

    # Se necessário, reposiciona o mouse para o local correto (aqui usa as mesmas coordenadas definidas)
    # time.sleep(2)
    pyautogui.moveTo(x=x, y=y)
    time.sleep(1)
    pyautogui.click(x=x, y=y)

    # Espera e clica no botão 'ok'
    not_fond= wait_for_image('botoes/not_found_2.png.png', confidence=0.9, timeout=3)
    print("asasasas")
    if not_fond is not  None:
        pyautogui.click(x=48, y=315)
        with open("nao_encontrados_3.txt", "a") as f:
            f.write(f"{contrato}\n")
        continue 
    print("ok")
    ok = wait_for_image('botoes/ok_2.png.png', confidence=0.9)
    print("ok encontrado")
    if ok is None:
        pyautogui.click(x=48, y=315)
        continue 
    pyautogui.click(ok)


    
    
    # Antes de clicar em "salvar", captura a data de modificação atual do arquivo de entrada

    
    # Espera e clica no botão 'salvar'
 
    if not monitorar_alteracao(original):
        continue
    salvar = wait_for_image('botoes/salvar_2.png.png', confidence=0.9)
    print("salvar encontrado")
    if salvar is None:
        continue 
    time.sleep(1)  # Pequena pausa para garantir que o botão esteja ativo
    pyautogui.click(salvar)
    pyautogui.click(salvar)


    # Em vez de usar um timer fixo, aguarda até que 'Dashboard.xlsx' seja atualizado
    time.sleep(1)  # Pequena pausa para garantir que o arquivo esteja acessível
    # Aguarda até que o arquivo 'Dashboard.xlsx' seja atualizado
    timeout = 60  # tempo máximo para aguardar, em segundos
    for _ in range(timeout):
        try:
            new_mod_time = os.path.getmtime(input_file)
        except Exception as e:
            print(f"Erro ao acessar 'Dashboard.xlsx': {e}")
            break

        if new_mod_time > old_mod_time:
            print("O arquivo 'Dashboard.xlsx' foi atualizado.")
            break
        time.sleep(1)

    print("Chamando salve_file")
    salve_file(str(contrato))
  