import glob
import io
import numpy as np
import os
import pandas as pd
import PyPDF2
import re
import shutil
import sys
import traceback
import threading
import warnings
import win32com.client as win32

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.cell.cell import MergedCell
from pandas import Timestamp
from pathlib import Path
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from tqdm import tqdm


def xlsx_to_pdf_one_page(input_file:str, output_path:str):


    os.makedirs(output_path,exist_ok=True)
    filename = os.path.splitext(os.path.basename(input_file))[0] + ".pdf"
    
    output_file = Path(output_path, filename)
    
    if os.path.isfile(output_file):
        return
    
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False

    try:
        workbook = excel.Workbooks.Open(input_file)
    except Exception as e:
        print(f"Erro ao abrir o arquivo {input_file}: {e}")
        return

    worksheet = workbook.Worksheets(1)  # planilha COM
    worksheet.Columns.AutoFit()

    # área de impressão de C3 até última coluna/linha
    last_column = 12
    last_row = 1
    for row in range(1, worksheet.UsedRange.Rows.Count + 1):
        borders = worksheet.Cells(row, 4).Borders
        for border_id in range(5, 13):
            if borders(border_id).Color != 0.0 or borders(border_id).LineStyle != -4142:
                last_row = row
                break    

    for column_index in range(1, worksheet.UsedRange.Columns.Count + 1):
        borders = worksheet.Cells(5, column_index).Borders
        for border_id in range(5, 13):
            if borders(border_id).Color != 0.0 or borders(border_id).LineStyle != -4142:
                # last_column = col
                break

    def column_index_to_letter(index):
        letter = ''
        while index > 0:
            index, rem = divmod(index - 1, 26)
            letter = chr(65 + rem) + letter
        return letter

    last_col_letter = column_index_to_letter(last_column)
    worksheet.PageSetup.PrintArea = worksheet.Range(f"C3:{last_col_letter}{last_row}").Address

    # configura impressão em 1 página
    worksheet.PageSetup.Zoom = False
    worksheet.PageSetup.FitToPagesWide = 1
    worksheet.PageSetup.FitToPagesTall = 1


    # **Auto-ajusta a largura de todas as colunas usadas**

    # exporta pra PDF
    workbook.ExportAsFixedFormat(0, str(output_file))

    workbook.Close(False)
    excel.Quit()
    #print(f"\nArquivo PDF gerado: {output_file}")