from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import unicodedata
import re
import time
from datetime import timedelta
import csv
import os
import zipfile  # <-- para capturar BadZipFile

from tqdm import tqdm  # pip install tqdm

TARGET_TEXT = "Diferença na Base de Cálculo entre IR e CS"


def normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))  # remove acentos
    s = re.sub(r"\s+", " ", s)  # colapsa espaços
    return s


TARGET_NORM = normalize_text(TARGET_TEXT)


def is_empty(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def last_value_in_row(ws, row_idx: int):
    # varre da última coluna usada para trás e pega o último valor não vazio
    for col_idx in range(ws.max_column, 0, -1):
        v = ws.cell(row=row_idx, column=col_idx).value
        if not is_empty(v):
            return v
    return None


def process_one_file(xlsx_path: Path) -> dict:
    """
    Processa um único arquivo XLSX e retorna o resultado.
    Importante: não deixa exceções "subirem" — retorna status de erro e segue o processamento.
    """
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)  # read_only acelera e reduz RAM
    except zipfile.BadZipFile:
        return {
            "arquivo": xlsx_path.name,
            "arquivo_completo": str(xlsx_path),
            "aba": None,
            "valor_ultimo_na_linha": None,
            "status": "erro: BadZipFile (xlsx inválido/corrompido)",
        }
    except PermissionError:
        return {
            "arquivo": xlsx_path.name,
            "arquivo_completo": str(xlsx_path),
            "aba": None,
            "valor_ultimo_na_linha": None,
            "status": "erro: PermissionError (sem acesso ao arquivo)",
        }
    except Exception as e:
        return {
            "arquivo": xlsx_path.name,
            "arquivo_completo": str(xlsx_path),
            "aba": None,
            "valor_ultimo_na_linha": None,
            "status": f"erro: {type(e).__name__} (falha ao abrir)",
        }

    try:
        for ws in wb.worksheets:
            for row_idx in range(1, ws.max_row + 1):
                b_val = ws.cell(row=row_idx, column=2).value  # coluna B
                if normalize_text(b_val) == TARGET_NORM:
                    ultimo = last_value_in_row(ws, row_idx)
                    return {
                        "arquivo": xlsx_path.name,
                        "arquivo_completo": str(xlsx_path),
                        "aba": ws.title,
                        "valor_ultimo_na_linha": ultimo,
                        "status": "ok",
                    }

        return {
            "arquivo": xlsx_path.name,
            "arquivo_completo": str(xlsx_path),
            "aba": None,
            "valor_ultimo_na_linha": None,
            "status": "nao_encontrado",
        }
    except Exception as e:
        return {
            "arquivo": xlsx_path.name,
            "arquivo_completo": str(xlsx_path),
            "aba": None,
            "valor_ultimo_na_linha": None,
            "status": f"erro: {type(e).__name__} (falha ao ler)",
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


class RunTimer:
    """Timer simples usando perf_counter, com formatação HH:MM:SS."""
    def __init__(self):
        self.t0 = None
        self.t1 = None

    def __enter__(self):
        self.t0 = time.perf_counter()
        return self

    def __exit__(self, exc_type, exc, tb):
        self.t1 = time.perf_counter()

    @property
    def elapsed_seconds(self) -> float:
        if self.t0 is None:
            return 0.0
        end = self.t1 if self.t1 is not None else time.perf_counter()
        return end - self.t0

    def fmt(self) -> str:
        return str(timedelta(seconds=int(self.elapsed_seconds)))


def human_rate(done: int, seconds: float) -> str:
    if seconds <= 0 or done <= 0:
        return "n/a"
    r = done / seconds
    return f"{r:.2f} arquivos/s ({r*60:.0f} arquivos/min)"


CHECKPOINT_PATH = Path("checkpoint.csv")
RESULT_PATH = Path("resultado.csv")
CHECKPOINT_FIELDS = ["arquivo", "arquivo_completo", "aba", "valor_ultimo_na_linha", "status"]


def load_checkpoint_processed_set(checkpoint_path: Path) -> set:
    """
    Lê o checkpoint.csv (se existir) e retorna um set com nomes de arquivos já processados.
    """
    if not checkpoint_path.exists():
        return set()

    processed = set()
    try:
        df_ck = pd.read_csv(checkpoint_path, dtype={"arquivo": str}, encoding="utf-8-sig")
        if "arquivo" in df_ck.columns:
            processed = set(df_ck["arquivo"].dropna().astype(str).tolist())
    except Exception:
        # fallback: leitura simples
        with checkpoint_path.open("r", newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("arquivo"):
                    processed.add(row["arquivo"])
    return processed


def ensure_checkpoint_header(checkpoint_path: Path):
    """
    Garante que o checkpoint.csv exista e tenha cabeçalho.
    """
    if not checkpoint_path.exists():
        with checkpoint_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=CHECKPOINT_FIELDS)
            writer.writeheader()
            f.flush()
            os.fsync(f.fileno())


def append_checkpoint_rows(checkpoint_path: Path, rows: list[dict]):
    """
    Append em lote no checkpoint.csv e força flush + fsync para reduzir perda em queda de energia.
    """
    if not rows:
        return

    with checkpoint_path.open("a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CHECKPOINT_FIELDS)

        for r in rows:
            writer.writerow({
                "arquivo": r.get("arquivo"),
                "arquivo_completo": r.get("arquivo_completo"),
                "aba": r.get("aba"),
                "valor_ultimo_na_linha": r.get("valor_ultimo_na_linha"),
                "status": r.get("status"),
            })

        f.flush()
        os.fsync(f.fileno())


def main():
    data_dir = Path(r"C:\Users\jaquelinerufino\Downloads\contratos_20260225\contratos_20260225")
    if not data_dir.exists() or not data_dir.is_dir():
        raise FileNotFoundError("Pasta 'data' não encontrada. Crie ./data e coloque os .xlsx lá.")

    files = sorted(data_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError("Nenhum arquivo .xlsx encontrado em ./data")

    # Checkpoint / retomada
    ensure_checkpoint_header(CHECKPOINT_PATH)
    processed = load_checkpoint_processed_set(CHECKPOINT_PATH)

    files_to_process = [p for p in files if p.name not in processed]

    print(f"Total .xlsx encontrados: {len(files)}")
    print(f"Já processados (checkpoint): {len(processed)}")
    print(f"Restantes para processar: {len(files_to_process)}")

    if not files_to_process:
        print("Nada a processar. Vou apenas gerar resultado.csv a partir do checkpoint.csv.")
        df_ck = pd.read_csv(CHECKPOINT_PATH, encoding="utf-8-sig")
        df_ck[["arquivo", "valor_ultimo_na_linha", "status"]].to_csv(RESULT_PATH, index=False, encoding="utf-8-sig")
        print(f"OK! Gerado: {RESULT_PATH.resolve()}")
        return

    # Batch de escrita do checkpoint (equilíbrio entre performance e segurança)
    CHECKPOINT_FLUSH_EVERY = 200  # grave a cada 200 resultados (ajuste se quiser)

    buffer_rows = []
    newly_processed = 0

    with RunTimer() as rt:
        for p in tqdm(files_to_process, total=len(files_to_process), desc="Processando XLSX", unit="arquivo"):
            r = process_one_file(p)
            buffer_rows.append(r)
            newly_processed += 1

            if len(buffer_rows) >= CHECKPOINT_FLUSH_EVERY:
                append_checkpoint_rows(CHECKPOINT_PATH, buffer_rows)
                buffer_rows.clear()

        # grava o que sobrou
        append_checkpoint_rows(CHECKPOINT_PATH, buffer_rows)
        buffer_rows.clear()

    elapsed = rt.elapsed_seconds

    # Resultado final: sempre construído a partir do checkpoint
    df_all = pd.read_csv(CHECKPOINT_PATH, encoding="utf-8-sig")
    df_all[["arquivo", "valor_ultimo_na_linha", "status"]].to_csv(RESULT_PATH, index=False, encoding="utf-8-sig")

    ok = int((df_all["status"] == "ok").sum())
    nao = int((df_all["status"] == "nao_encontrado").sum())
    err = int((df_all["status"].astype(str).str.startswith("erro")).sum())

    print(f"\nOK! Gerado: {RESULT_PATH.resolve()}")
    print(df_all["status"].value_counts(dropna=False))

    print("\n===== RESUMO DA EXECUÇÃO =====")
    print(f"Arquivos processados nesta execução: {newly_processed}")
    print(f"Total acumulado no checkpoint: {len(df_all)}")
    print(f"OK: {ok} | Não encontrado: {nao} | Erro: {err}")
    print(f"Tempo desta execução: {rt.fmt()} ({elapsed:.2f}s)")
    print(f"Throughput médio (nesta execução): {human_rate(max(newly_processed, 1), elapsed)}")
    print("==============================\n")


if __name__ == "__main__":
    main()